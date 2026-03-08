#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Standalone Σύστημα Πιστοποιητικών Υγείας
Απομονωμένο από το κύριο ιατρικό πρόγραμμα
"""

import sys
import os
import uuid
import re
import sqlite3  # 🔽 ΠΡΟΣΘΗΚΗ
from datetime import datetime, date  # 🔽 ΒΕΒΑΙΩΣΗ ότι υπάρχει
from decimal import Decimal

from PyQt5.QtWidgets import (
        QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
        QLabel, QLineEdit, QTextEdit, QPushButton, QDateEdit,
        QComboBox, QTreeWidget, QTreeWidgetItem, QTabWidget,
        QMessageBox, QFileDialog, QDialog, QListWidget,
        QAbstractItemView, QHeaderView, QInputDialog,
        QGroupBox, QStatusBar, QCheckBox  # 🔽 ΠΡΟΣΘΗΚΗ
    )
from PyQt5.QtCore import Qt, QDate, QRegExp, Qt, QTimer
from PyQt5.QtGui import QRegExpValidator, QFont
from database import Database, Config

class HealthCertificateApp(QMainWindow):
    """Κύριο παράθυρο εφαρμογής πιστοποιητικών υγείας."""

    def __init__(self):
        super().__init__()

        self.page_size = 100
        self.current_offset = 0

        self.current_user_id = None
        self.current_username = None
        self.current_user_role = None

        self.setStyleSheet(self.get_stylesheet())

        self.statusBar().showMessage("Εκκίνηση εφαρμογής...")

        self.setWindowTitle(f"{Config.APP_NAME} v{Config.APP_VERSION}")
        self.setMinimumSize(1000, 700)
        self.resize(1200, 800)
        
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.main_layout = QVBoxLayout()
        self.central_widget.setLayout(self.main_layout)

        self.setup_login_bar()

        self.tabs = QTabWidget()
        self.main_layout.addWidget(self.tabs)
        self.tabs.currentChanged.connect(self.on_tab_changed)

        self.setup_certificates_tab()

        QTimer.singleShot(0, self.initialize_system)

    def on_tab_changed(self, index):

        tab_name = self.tabs.tabText(index)

        if tab_name == "Πιστοποιητικά Υγείας":
            self.load_certificates()
            
    def initialize_system(self):

        db_path = os.path.join(
            os.path.dirname(__file__),
            "health_certificates.db"
        )

        if not os.path.exists(db_path):

            if not Database.create_tables():
                QMessageBox.critical(self, "Σφάλμα", "Αδυναμία δημιουργίας βάσης δεδομένων.")
                sys.exit(1)

        Database.create_default_admin()
        Database.setup_directories()
        Database.setup_backup_directory()

        self.statusBar().showMessage("Έτοιμο")

        self.show_login_dialog()

    def get_stylesheet(self):
        """Επιστρέφει το stylesheet για την εφαρμογή."""
        return """
        /* Γενικά στυλ */
        QMainWindow {
            background-color: #f0f4f8;
        }

        QWidget {
            font-family: 'Segoe UI', Arial, sans-serif;
            font-size: 12px;
        }

        /* Labels */
        QLabel {
            color: #2c3e50;
            font-weight: 500;
        }

        QLabel[title="true"] {
            font-size: 14px;
            font-weight: bold;
            color: #2980b9;
            padding: 5px;
            border-bottom: 2px solid #3498db;
        }

        /* Input fields */
        QLineEdit, QTextEdit, QDateEdit, QComboBox {
            background-color: white;
            border: 1px solid #bdc3c7;
            border-radius: 4px;
            padding: 6px;
            min-height: 25px;
        }

        QLineEdit:focus, QTextEdit:focus, QDateEdit:focus, QComboBox:focus {
            border: 2px solid #3498db;
            background-color: #e8f4fc;
        }

        QLineEdit[error="true"] {
            border: 2px solid #e74c3c;
            background-color: #ffeaea;
        }

        /* Buttons */
        QPushButton {
            background-color: #3498db;
            color: white;
            border: none;
            border-radius: 4px;
            padding: 8px 16px;
            font-weight: 600;
            min-height: 30px;
        }

        QPushButton:hover {
            background-color: #2980b9;
        }

        QPushButton:pressed {
            background-color: #1c598a;
        }

        QPushButton[role="add"] {
            background-color: #27ae60;
        }

        QPushButton[role="add"]:hover {
            background-color: #219653;
        }

        QPushButton[role="delete"] {
            background-color: #e74c3c;
        }

        QPushButton[role="delete"]:hover {
            background-color: #c0392b;
        }

        QPushButton[role="update"] {
            background-color: #f39c12;
        }

        QPushButton[role="update"]:hover {
            background-color: #d68910;
        }

        QPushButton[role="export"] {
            background-color: #9b59b6;
        }

        QPushButton[role="export"]:hover {
            background-color: #8e44ad;
        }

        QPushButton[role="attachment"] {
            background-color: #1abc9c;
        }

        QPushButton[role="attachment"]:hover {
            background-color: #16a085;
        }

        /* Tree Widget */
        QTreeWidget {
            background-color: white;
            border: 1px solid #bdc3c7;
            border-radius: 4px;
            alternate-background-color: #f8f9fa;
        }

        QTreeWidget::item {
            padding: 6px;
            border-bottom: 1px solid #ecf0f1;
        }

        QTreeWidget::item:selected {
            background-color: #3498db;
            color: white;
        }

        QTreeWidget::item:hover {
            background-color: #e8f4fc;
        }

        QHeaderView::section {
            background-color: #2c3e50;
            color: white;
            padding: 8px;
            border: none;
            font-weight: bold;
        }

        /* Tabs */
        QTabWidget::pane {
            border: 1px solid #bdc3c7;
            background-color: white;
            border-radius: 4px;
        }

        QTabBar::tab {
            background-color: #ecf0f1;
            color: #2c3e50;
            padding: 8px 16px;
            margin-right: 2px;
            border-top-left-radius: 4px;
            border-top-right-radius: 4px;
        }

        QTabBar::tab:selected {
            background-color: #3498db;
            color: white;
            font-weight: bold;
        }

        QTabBar::tab:hover {
            background-color: #d5dbdb;
        }

        /* List Widget */
        QListWidget {
            background-color: white;
            border: 1px solid #bdc3c7;
            border-radius: 4px;
            padding: 4px;
        }

        QListWidget::item {
            padding: 6px;
            border-bottom: 1px solid #ecf0f1;
        }

        QListWidget::item:selected {
            background-color: #3498db;
            color: white;
        }

        /* Dialog */
        QDialog {
            background-color: white;
        }

        /* Login bar */
        #login_bar {
            background-color: #2c3e50;
            padding: 8px;
            border-bottom: 3px solid #3498db;
        }

        #user_label {
            color: white;
            font-weight: bold;
            font-size: 13px;
        }

        /* Status bar */
        QStatusBar {
            background-color: #2c3e50;
            color: white;
        }

        /* Scrollbars */
        QScrollBar:vertical {
            background-color: #ecf0f1;
            width: 12px;
        }

        QScrollBar::handle:vertical {
            background-color: #95a5a6;
            border-radius: 6px;
        }

        QScrollBar::handle:vertical:hover {
            background-color: #7f8c8d;
        }

        /* Group boxes */
        QGroupBox {
            border: 2px solid #3498db;
            border-radius: 6px;
            margin-top: 10px;
            font-weight: bold;
            color: #2980b9;
            padding-top: 10px;
        }

        QGroupBox::title {
            subcontrol-origin: margin;
            left: 10px;
            padding: 0 5px 0 5px;
        }
        """

    def setup_login_bar(self):
        """Ρυθμίζει τη γραμμή login/logout με χρώματα."""
        login_bar = QWidget()
        login_bar.setObjectName("login_bar")
        login_bar.setStyleSheet("""
            QWidget#login_bar {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #2c3e50, stop:1 #3498db);
                border-bottom: 3px solid #2980b9;
            }
        """)

        login_layout = QHBoxLayout(login_bar)
        login_layout.setContentsMargins(10, 5, 10, 5)

        # Τίτλος εφαρμογής
        title_label = QLabel(f"🏥 {Config.APP_NAME}")
        title_label.setStyleSheet("""
            QLabel {
                color: white;
                font-size: 16px;
                font-weight: bold;
                font-family: 'Segoe UI', Arial;
            }
        """)

        # Χρήστης
        self.user_label = QLabel("👤 Παρακαλώ συνδεθείτε")
        self.user_label.setObjectName("user_label")
        self.user_label.setStyleSheet("""
            QLabel#user_label {
                color: white;
                font-size: 13px;
                background-color: rgba(255, 255, 255, 0.2);
                padding: 5px 10px;
                border-radius: 3px;
            }
        """)

        # 🔽 ΝΕΟ ΚΟΥΜΠΙ: Διαχείριση Χρηστών (μόνο για διαχειριστές)
        self.manage_users_btn = QPushButton("👤 Διαχείριση Χρηστών")
        self.manage_users_btn.setStyleSheet("""
            QPushButton {
                background-color: #9b59b6;
                color: white;
                border: none;
                border-radius: 3px;
                padding: 6px 12px;
                font-weight: bold;
                margin-right: 5px;
            }
            QPushButton:hover {
                background-color: #8e44ad;
            }
            QPushButton:disabled {
                background-color: #7d3c98;
                color: #cccccc;
            }
        """)
        self.manage_users_btn.clicked.connect(self.manage_users)
        self.manage_users_btn.hide()  # Αρχικά κρυφό

        # 🔽 ΠΡΟΣΘΗΚΗ: Κουμπί Ρυθμίσεων
        self.settings_btn = QPushButton("⚙️ Ρυθμίσεις")
        self.settings_btn.setStyleSheet("""
            QPushButton {
                background-color: #f39c12;
                color: white;
                border: none;
                border-radius: 3px;
                padding: 6px 12px;
                font-weight: bold;
                margin-right: 5px;
            }
            QPushButton:hover {
                background-color: #d68910;
            }
        """)
        self.settings_btn.clicked.connect(self.show_settings_dialog)

        self.logout_btn = QPushButton("🚪 Αποσύνδεση")
        self.logout_btn.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                border-radius: 3px;
                padding: 6px 12px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
        """)
        self.logout_btn.clicked.connect(self.logout)
        self.logout_btn.hide()

        # Προσθήκη widgets στο layout
        login_layout.addWidget(title_label)
        login_layout.addStretch()
        login_layout.addWidget(self.user_label)
        login_layout.addWidget(self.manage_users_btn)
        login_layout.addWidget(self.settings_btn)  # 🔽 ΠΡΟΣΘΗΚΗ: Ρυθμίσεις
        login_layout.addWidget(self.logout_btn)

        self.main_layout.addWidget(login_bar)

    def show_settings_dialog(self):
        """Εμφανίζει διάλογο ρυθμίσεων."""
        dialog = QDialog(self)
        dialog.setWindowTitle("⚙️ Ρυθμίσεις Εφαρμογής")
        dialog.setMinimumWidth(400)

        layout = QVBoxLayout(dialog)

        # Τίτλος
        title_label = QLabel("Ρυθμίσεις Εφαρμογής")
        title_label.setStyleSheet("""
            QLabel {
                font-size: 16px;
                font-weight: bold;
                color: #2980b9;
                padding-bottom: 10px;
                border-bottom: 2px solid #3498db;
            }
        """)
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)

        # Ρυθμίσεις backup
        backup_group = QGroupBox("💾 Ρυθμίσεις Backup")
        backup_layout = QVBoxLayout()

        self.auto_backup_checkbox = QCheckBox("Αυτόματο backup κατά το κλείσιμο")
        self.auto_backup_checkbox.setChecked(True)  # Ενεργό από default

        self.backup_notification_checkbox = QCheckBox("Ειδοποίηση για backup")
        self.backup_notification_checkbox.setChecked(True)

        backup_layout.addWidget(self.auto_backup_checkbox)
        backup_layout.addWidget(self.backup_notification_checkbox)
        backup_group.setLayout(backup_layout)
        layout.addWidget(backup_group)

        # Κουμπιά
        buttons_layout = QHBoxLayout()

        save_btn = QPushButton("💾 Αποθήκευση")
        save_btn.clicked.connect(lambda: self.save_settings(dialog))

        cancel_btn = QPushButton("❌ Ακύρωση")
        cancel_btn.clicked.connect(dialog.reject)

        buttons_layout.addStretch()
        buttons_layout.addWidget(save_btn)
        buttons_layout.addWidget(cancel_btn)

        layout.addLayout(buttons_layout)

        # Φόρτωση τρέχουσες ρυθμίσεις
        self.load_settings()

        dialog.exec_()

    def load_settings(self):
        """Φορτώνει τις ρυθμίσεις από αρχείο."""
        try:
            import json
            import os

            settings_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'settings.json')

            if os.path.exists(settings_file):
                with open(settings_file, 'r', encoding='utf-8') as f:
                    settings = json.load(f)

                    # Ενημέρωση checkboxes αν υπάρχουν
                    if hasattr(self, 'auto_backup_checkbox'):
                        self.auto_backup_checkbox.setChecked(settings.get('auto_backup', True))
                    if hasattr(self, 'backup_notification_checkbox'):
                        self.backup_notification_checkbox.setChecked(settings.get('backup_notification', True))
        except Exception as e:
            print(f"Σφάλμα φόρτωσης ρυθμίσεων: {e}")

    def save_settings(self, dialog):
        """Αποθηκεύει τις ρυθμίσεις σε αρχείο."""
        try:
            import json
            import os

            settings = {
                'auto_backup': self.auto_backup_checkbox.isChecked(),
                'backup_notification': self.backup_notification_checkbox.isChecked()
            }

            settings_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'settings.json')

            with open(settings_file, 'w', encoding='utf-8') as f:
                json.dump(settings, f, indent=4, ensure_ascii=False)

            QMessageBox.information(dialog, "Επιτυχία", "Οι ρυθμίσεις αποθηκεύτηκαν.")
            dialog.accept()

        except Exception as e:
            QMessageBox.critical(dialog, "Σφάλμα", f"Σφάλμα αποθήκευσης ρυθμίσεων: {e}")
            
    def closeEvent(self, event):
        """
        Εκτελείται όταν κλείνει η εφαρμογή.
        Δημιουργεί αυτόματο backup αν υπάρχουν δεδομένα.
        """
        try:
            # Έλεγχος ρυθμίσεων
            import json
            import os

            settings_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'settings.json')
            auto_backup_enabled = True  # Default τιμή

            if os.path.exists(settings_file):
                try:
                    with open(settings_file, 'r', encoding='utf-8') as f:
                        settings = json.load(f)
                        auto_backup_enabled = settings.get('auto_backup', True)
                except:
                    pass  # Χρησιμοποιούμε default τιμή

            # Αν το auto backup είναι απενεργοποιημένο, κλείνουμε κανονικά
            if not auto_backup_enabled:
                event.accept()
                return

            # Έλεγχος αν υπάρχουν πιστοποιητικά για backup
            if self.current_user_id and self.certificates_table.topLevelItemCount() > 0:
                # Έλεγχος ειδοποιήσεων
                show_notification = True
                if os.path.exists(settings_file):
                    try:
                        with open(settings_file, 'r', encoding='utf-8') as f:
                            settings = json.load(f)
                            show_notification = settings.get('backup_notification', True)
                    except:
                        pass

                if show_notification:
                    reply = QMessageBox.question(
                        self, "Αυτόματο Backup",
                        "Θέλετε να δημιουργηθεί backup των δεδομένων πριν το κλείσιμο;\n\n"
                        "⚠️  Συνιστάται για ασφάλεια δεδομένων.",
                        QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel,
                        QMessageBox.Yes
                    )

                    if reply == QMessageBox.Yes:
                        # Δημιουργία backup
                        success = self.create_auto_backup()
                        if not success:
                            reply2 = QMessageBox.question(
                                self, "Σφάλμα Backup",
                                "Το backup απέτυχε. Θέλετε να κλείσετε την εφαρμογή χωρίς backup;",
                                QMessageBox.Yes | QMessageBox.No,
                                QMessageBox.No
                            )
                            if reply2 != QMessageBox.Yes:
                                event.ignore()
                                return
                    elif reply == QMessageBox.Cancel:
                        # Ακύρωση κλεισίματος
                        event.ignore()
                        return
                    # Αν απαντήσει No, συνεχίζουμε χωρίς backup
                else:
                    # Αυτόματο backup χωρίς ερώτηση
                    self.create_auto_backup()

        except Exception as e:
            print(f"Σφάλμα κατά το κλείσιμο: {e}")
            # Σε περίπτωση σφάλματος, απλά κλείνουμε
            pass

        # Συνέχιση με το κανονικό κλείσιμο
        event.accept()

    def create_auto_backup(self):
        """
        Δημιουργεί αυτόματο backup σε CSV αρχείο.
        Επιστρέφει True αν ήταν επιτυχές, False αν απέτυχε.
        """
        try:
            import os
            import csv
            from datetime import datetime

            if self.certificates_table.topLevelItemCount() == 0:
                return True  # Δεν υπάρχουν δεδομένα, αλλά δεν είναι σφάλμα

            # Δημιουργία backup φακέλου αν δεν υπάρχει
            backup_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'backups', 'auto_backup')
            os.makedirs(backup_dir, exist_ok=True)

            # Δημιουργία ονόματος αρχείου με ημερομηνία
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            username = self.current_username if self.current_username else 'unknown'
            username = username.replace(' ', '_')
            filename = f"auto_backup_{username}_{timestamp}.csv"
            file_path = os.path.join(backup_dir, filename)

            conn = Database.get_connection()
            if not conn:
                return False

            try:
                cur = conn.cursor()

                # Ανάκτηση όλων των δεδομένων του τρέχοντος χρήστη
                cur.execute("""
                    SELECT 
                        protocol_number,
                        hc_last_name,
                        hc_first_name,
                        hc_father_name,
                        hc_amka,
                        hc_personal_number,
                        id_number,
                        passport_number,
                        certificate_date,
                        residence,
                        work_type,
                        comments
                    FROM health_certificates
                    WHERE doctor_id = ?
                    ORDER BY certificate_date DESC
                """, (self.current_user_id,))

                certificates = cur.fetchall()

                # Γράψιμο στο CSV
                with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                    writer = csv.writer(csvfile, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)

                    # Επικεφαλίδες
                    headers = [
                        'Πρωτόκολλο', 'Επώνυμο', 'Όνομα', 'Πατρώνυμο',
                        'ΑΜΚΑ', 'Προσωπικός Αριθμός', 'Ταυτότητα', 'Διαβατήριο',
                        'Ημερομηνία Πιστοποιητικού', 'Κατοικία', 'Είδος Εργασίας', 'Σχόλια'
                    ]
                    writer.writerow(headers)

                    # Δεδομένα
                    count = 0
                    for cert in certificates:
                        row = []
                        for i, value in enumerate(cert):
                            if value is None:
                                row.append('')
                            elif i == 8 and value:  # Ημερομηνία
                                try:
                                    # Μετατροπή ημερομηνίας σε ελληνική μορφή
                                    if isinstance(value, str):
                                        try:
                                            date_obj = datetime.strptime(value.split()[0], '%Y-%m-%d')
                                        except:
                                            date_obj = datetime.strptime(value, '%Y-%m-%d %H:%M:%S')
                                        row.append(date_obj.strftime('%d/%m/%Y'))
                                    else:
                                        # Αν είναι datetime object
                                        row.append(value.strftime('%d/%m/%Y'))
                                except:
                                    row.append(str(value))
                            else:
                                row.append(str(value))
                        writer.writerow(row)
                        count += 1

                # Εγγραφή σε log αρχείο
                self.log_backup_action(file_path, count)

                return True

            except Exception as e:
                print(f"Σφάλμα κατά τη δημιουργία backup: {e}")
                return False

        except Exception as e:
            print(f"Σφάλμα στη μέθοδο create_auto_backup: {e}")
            return False

    def log_backup_action(self, file_path, record_count):
        """Κρατάει log των backup που έχουν γίνει."""
        try:
            import os
            from datetime import datetime

            log_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'backups', 'logs')
            os.makedirs(log_dir, exist_ok=True)

            log_file = os.path.join(log_dir, 'backup_log.csv')

            # Ελέγχουμε αν υπάρχει το log αρχείο για να προσθέσουμε headers
            file_exists = os.path.exists(log_file)

            with open(log_file, 'a', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)

                if not file_exists:
                    writer.writerow(['Ημερομηνία', 'Χρήστης', 'Αρχείο', 'Εγγραφές', 'Τύπος'])

                writer.writerow([
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    self.current_username or 'unknown',
                    os.path.basename(file_path),
                    record_count,
                    'Αυτόματο (Κλείσιμο)'
                ])

        except Exception as e:
            print(f"Σφάλμα κατά την εγγραφή log: {e}")

    def show_login_dialog(self):
        """Εμφανίζει διάλογο σύνδεσης με ταυτόχρονη εισαγωγή username & password."""

        # Δημιουργία custom dialog
        dialog = QDialog(self)
        dialog.setWindowTitle("🔐 Σύνδεση στο Σύστημα")
        dialog.setMinimumWidth(400)
        dialog.setStyleSheet("""
            QDialog {
                background-color: #f8f9fa;
            }
            QLabel {
                font-weight: bold;
                color: #2c3e50;
            }
            QLineEdit {
                padding: 8px;
                border: 1px solid #ddd;
                border-radius: 4px;
                font-size: 14px;
            }
            QLineEdit:focus {
                border: 2px solid #3498db;
            }
            QPushButton {
                padding: 10px 20px;
                border-radius: 4px;
                font-weight: bold;
                font-size: 14px;
                min-width: 100px;
            }
        """)

        layout = QVBoxLayout(dialog)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)

        # Τίτλος
        title_label = QLabel("Σύνδεση στο Σύστημα Πιστοποιητικών Υγείας")
        title_label.setStyleSheet("""
            QLabel {
                font-size: 16px;
                font-weight: bold;
                color: #2980b9;
                padding-bottom: 10px;
                border-bottom: 2px solid #3498db;
            }
        """)
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)

        # Username
        username_layout = QVBoxLayout()
        username_label = QLabel("👤 Όνομα Χρήστη:")
        self.login_username_input = QLineEdit()
        self.login_username_input.setPlaceholderText("Εισάγετε το όνομα χρήστη")
        self.login_username_input.setText("p_ygeias")  # Προεπιλεγμένο για ευκολία

        username_layout.addWidget(username_label)
        username_layout.addWidget(self.login_username_input)
        layout.addLayout(username_layout)

        # Password
        password_layout = QVBoxLayout()
        password_label = QLabel("🔑 Κωδικός:")
        self.login_password_input = QLineEdit()
        self.login_password_input.setPlaceholderText("Εισάγετε τον κωδικό")
        self.login_password_input.setEchoMode(QLineEdit.Password)
        self.login_password_input.setText("")  # Προεπιλεγμένο για ευκολία

        password_layout.addWidget(password_label)
        password_layout.addWidget(self.login_password_input)
        layout.addLayout(password_layout)

        # Προεπιλεγμένοι λογαριασμοί (γρήγορη επιλογή)
        accounts_group = QGroupBox("📋 Προεπιλεγμένοι Λογαριασμοί")
        accounts_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                border: 1px solid #bdc3c7;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
                color: #7f8c8d;
            }
        """)

        accounts_layout = QVBoxLayout()

        # Doctor button
        doctor_btn = QPushButton("👤 Γιατρός:")
        doctor_btn.setStyleSheet("""
            QPushButton {
                text-align: left;
                padding: 8px;
                background-color: #e8f4fc;
                border: 1px solid #3498db;
            }
            QPushButton:hover {
                background-color: #d4e6f1;
            }
        """)
        #doctor_btn.clicked.connect(lambda: self.fill_login_credentials("p_ygeias", "1234"))

        # Admin button
        admin_btn = QPushButton("👑 Διαχειριστής: admin")
        admin_btn.setStyleSheet("""
            QPushButton {
                text-align: left;
                padding: 8px;
                background-color: #f4ecf7;
                border: 1px solid #9b59b6;
            }
            QPushButton:hover {
                background-color: #e8daef;
            }
        """)
        admin_btn.clicked.connect(lambda: self.fill_login_credentials("admin", "admin123"))

        accounts_layout.addWidget(doctor_btn)
        accounts_layout.addWidget(admin_btn)
        accounts_group.setLayout(accounts_layout)
        layout.addWidget(accounts_group)

        # Κουμπιά
        buttons_layout = QHBoxLayout()

        login_btn = QPushButton("✅ Σύνδεση")
        login_btn.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
            }
            QPushButton:hover {
                background-color: #219653;
            }
            QPushButton:pressed {
                background-color: #1e8449;
            }
        """)
        login_btn.clicked.connect(lambda: self.process_login(dialog))

        cancel_btn = QPushButton("❌ Ακύρωση")
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
        """)
        cancel_btn.clicked.connect(lambda: self.cancel_login(dialog))

        buttons_layout.addWidget(login_btn)
        buttons_layout.addWidget(cancel_btn)
        layout.addLayout(buttons_layout)

        # Enter key για γρήγορη σύνδεση
        self.login_username_input.returnPressed.connect(lambda: self.login_password_input.setFocus())
        self.login_password_input.returnPressed.connect(lambda: self.process_login(dialog))

        # Εστίαση στο username
        self.login_username_input.setFocus()

        # Εκτέλεση του dialog
        dialog.exec_()

    def fill_login_credentials(self, username, password):
        """Συμπληρώνει αυτόματα τα στοιχεία σύνδεσης."""
        self.login_username_input.setText(username)
        self.login_password_input.setText(password)
        self.login_password_input.setFocus()

    def process_login(self, dialog):
        """Επεξεργάζεται την προσπάθεια σύνδεσης."""
        username = self.login_username_input.text().strip()
        password = self.login_password_input.text().strip()

        if not username or not password:
            QMessageBox.warning(dialog, "Σφάλμα",
                                "Παρακαλώ συμπληρώστε και τα δύο πεδία.")
            return

        if self.authenticate_user(username, password):
            self.user_label.setText(f"Συνδεδεμένος ως: {username}")
            self.logout_btn.show()
            self.load_certificates()
            self.refresh_protocol()
            dialog.accept()
        else:
            QMessageBox.warning(dialog, "Σφάλμα",
                                "Λάθος όνομα χρήστη ή κωδικός.\n\n"
                                "Δοκιμάστε:\n"
                                "• Διαχειριστής: admin / admin123"
                                # Ή:
                                "Επικοινωνήστε με τον διαχειριστή για λογαριασμό")
            self.login_password_input.clear()
            self.login_password_input.setFocus()

    def cancel_login(self, dialog):
        """Ακυρώνει τη σύνδεση."""
        reply = QMessageBox.question(
            dialog, "Έξοδος",
            "Θέλετε να κλείσετε το πρόγραμμα;",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            dialog.reject()
            sys.exit(0)
        else:
            # Μείνετε στο login dialog
            self.login_username_input.setFocus()

    def generate_protocol_number(self):
        """Δημιουργεί αυτόματα τον αριθμό πρωτοκόλλου με βάση το τελευταίο αριθμό για το τρέχον έτος."""
        from datetime import datetime

        if not self.current_user_id:
            current_year = datetime.now().year
            return f"0001/01-01-{current_year}", 1, current_year

        current_year = datetime.now().year

        conn = Database.get_connection()
        if not conn:
            current_year = datetime.now().year
            return f"0001/01-01-{current_year}", 1, current_year

        try:
            cur = conn.cursor()

            # Βρες τον τελευταίο αριθμό πρωτοκόλλου για το τρέχον έτος
            cur.execute("""
                SELECT MAX(proto_seq) as max_seq
                FROM health_certificates
                WHERE doctor_id = ? AND proto_year = ?
            """, (self.current_user_id, current_year))

            result = cur.fetchone()

            if result and result[0] is not None:
                next_seq = int(result[0]) + 1
            else:
                next_seq = 1

            # Μορφοποίηση: ΝΝΝΝ/DD-MM-YYYY
            today = datetime.now()
            protocol_number = f"{next_seq:04d}/{today.day:02d}-{today.month:02d}-{today.year}"

            return protocol_number, next_seq, current_year

        except Exception as e:
            print(f"Σφάλμα δημιουργίας πρωτοκόλλου: {e}")
            import traceback
            traceback.print_exc()
            today = datetime.now()
            return f"0001/{today.day:02d}-{today.month:02d}-{today.year}", 1, current_year

    def change_protocol_date(self):
        """Αλλάζει την ημερομηνία του πρωτοκόλλου και ανανεώνει την αρίθμηση."""

        # Διάλογος για επιλογή νέας ημερομηνίας
        from PyQt5.QtWidgets import QDialog, QVBoxLayout, QLabel, QDateEdit, QPushButton, QHBoxLayout

        dialog = QDialog(self)
        dialog.setWindowTitle("📅 Αλλαγή Ημερομηνίας Πρωτοκόλλου")
        dialog.setMinimumWidth(300)

        layout = QVBoxLayout(dialog)

        # Οδηγίες
        label = QLabel(
            "Επιλέξτε νέα ημερομηνία για το πρωτόκολλο:\n\n"
            "Η αρίθμηση θα ξεκινήσει εκ νέου από 1\n"
            "για το επιλεγμένο έτος."
        )
        label.setWordWrap(True)
        layout.addWidget(label)

        # Ημερομηνία
        date_label = QLabel("Νέα Ημερομηνία:")
        self.new_protocol_date = QDateEdit()
        self.new_protocol_date.setDate(QDate.currentDate())
        self.new_protocol_date.setCalendarPopup(True)
        self.new_protocol_date.setDisplayFormat("dd/MM/yyyy")

        layout.addWidget(date_label)
        layout.addWidget(self.new_protocol_date)

        # Κουμπιά
        buttons_layout = QHBoxLayout()

        ok_btn = QPushButton("✅ Εφαρμογή")
        ok_btn.clicked.connect(lambda: self.apply_new_protocol_date(dialog))

        cancel_btn = QPushButton("❌ Ακύρωση")
        cancel_btn.clicked.connect(dialog.reject)

        buttons_layout.addWidget(ok_btn)
        buttons_layout.addWidget(cancel_btn)
        layout.addLayout(buttons_layout)

        dialog.exec_()

    def apply_new_protocol_date(self, dialog):
        """Εφαρμόζει τη νέα ημερομηνία πρωτοκόλλου."""
        new_date = self.new_protocol_date.date().toPyDate()

        # Δημιουργία νέου αριθμού πρωτοκόλλου με βάση τη νέα ημερομηνία
        protocol_data = self.generate_protocol_number_with_date(new_date)

        if isinstance(protocol_data, tuple):
            protocol_number, _, _ = protocol_data
        else:
            protocol_number = protocol_data

        # Ενημέρωση του πεδίου πρωτοκόλλου
        self.protocol_input.setText(protocol_number)

        QMessageBox.information(
            self, "Επιτυχία",
            f"Νέο πρωτόκολλο: {protocol_number}\n"
            f"Ημερομηνία: {new_date.strftime('%d/%m/%Y')}"
        )

        dialog.accept()

    def generate_protocol_number_with_date(self, target_date):
        """Δημιουργεί πρωτόκολλο με βάση συγκεκριμένη ημερομηνία."""
        from datetime import datetime

        if not self.current_user_id:
            return f"0001/{target_date.day:02d}-{target_date.month:02d}-{target_date.year}", 1, target_date.year

        conn = Database.get_connection()
        if not conn:
            return f"0001/{target_date.day:02d}-{target_date.month:02d}-{target_date.year}", 1, target_date.year

        try:
            cur = conn.cursor()

            # Βρες τον τελευταίο αριθμό πρωτοκόλλου για το συγκεκριμένο έτος
            cur.execute("""
                SELECT protocol_number, proto_seq
                FROM health_certificates
                WHERE doctor_id = ? AND proto_year = ?
                ORDER BY proto_seq DESC LIMIT 1
            """, (self.current_user_id, target_date.year))

            result = cur.fetchone()

            if result:
                next_seq = int(result[1]) + 1
            else:
                next_seq = 1  # Νέο έτος, ξεκινάμε από 1

            # Μορφοποίηση: ΝΝΝΝ/DD-MM-YYYY
            protocol_number = f"{next_seq:04d}/{target_date.day:02d}-{target_date.month:02d}-{target_date.year}"

            return protocol_number, next_seq, target_date.year

        except Exception as e:
            print(f"Σφάλμα δημιουργίας πρωτοκόλλου: {e}")
            import traceback
            traceback.print_exc()
            return f"0001/{target_date.day:02d}-{target_date.month:02d}-{target_date.year}", 1, target_date.year

    def authenticate_user(self, username, password):
        """Ελέγχει τα στοιχεία σύνδεσης."""
        conn = Database.get_connection()
        if not conn:
            return False

        try:
            cur = conn.cursor()

            cur.execute("""
                SELECT id, username, role FROM users 
                WHERE username = ? AND password_hash = ? AND is_active = TRUE
            """, (username, password))

            user = cur.fetchone()
            if user:
                self.current_user_id = user[0]
                self.current_username = user[1]
                self.current_user_role = user[2]

                # 🔽 ΕΝΗΜΕΡΩΣΗ: Εμφάνιση κουμπιού διαχείρισης μόνο για διαχειριστές
                if self.current_user_role == 'admin':
                    self.manage_users_btn.show()
                else:
                    self.manage_users_btn.hide()

                return True

        except Exception as e:
            print(f"Σφάλμα authentication: {e}")
            import traceback
            traceback.print_exc()
        return False

    def manage_users(self):
        """Ανοίγει το παράθυρο διαχείρισης χρηστών."""
        if self.current_user_role != 'admin':
            QMessageBox.warning(self, "Προσοχή",
                                "Μόνο οι διαχειριστές μπορούν να διαχειριστούν χρήστες.")
            return

        dialog = UserManagementDialog(self)
        dialog.exec_()
    
    def logout(self):
        """Αποσύνδεση χρήστη."""
        self.current_user_id = None
        self.current_username = None
        self.current_user_role = None
        
        self.user_label.setText("Παρακαλώ συνδεθείτε")
        self.logout_btn.hide()
        self.show_login_dialog()

    def setup_certificates_tab(self):
        """Ρυθμίζει την καρτέλα πιστοποιητικών."""
        tab = QWidget()
        tab.setObjectName("certificates_tab")
        self.tabs.addTab(tab, "📋 Πιστοποιητικά Υγείας")

        # Κύριο layout
        main_layout = QVBoxLayout()
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(10, 10, 10, 10)
        tab.setLayout(main_layout)

        # === ΟΜΑΔΑ ΦΟΡΜΑΣ ===
        form_group = QGroupBox("📝 Νέο Πιστοποιητικό")
        form_group.setStyleSheet("""
            QGroupBox {
                font-size: 13px;
                font-weight: bold;
                color: #2980b9;
            }
        """)
        form_layout = QVBoxLayout()
        form_group.setLayout(form_layout)

        # Προσθήκη χρώματος στα πεδία
        print("DEBUG: Αρχίζω setup_form_fields")  # Debug
        self.setup_form_fields(form_layout)  # Αυτό θα αλλάξουμε ξεχωριστά
        print("DEBUG: Τέλος setup_form_fields")  # Debug

        # Κουμπιά με χρώματα
        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(10)

        # ΚΟΥΜΠΙΑ ΜΕ ΧΡΩΜΑΤΑ ΚΑΙ ICONS
        self.add_btn = QPushButton("➕ Προσθήκη")
        self.add_btn.setProperty("role", "add")
        self.add_btn.setIcon(self.style().standardIcon(getattr(self.style(), 'SP_DialogApplyButton', None)))
        self.add_btn.clicked.connect(self.add_certificate)

        self.update_btn = QPushButton("✏️ Ενημέρωση")
        self.update_btn.setProperty("role", "update")
        self.update_btn.setIcon(self.style().standardIcon(getattr(self.style(), 'SP_BrowserReload', None)))
        self.update_btn.clicked.connect(self.update_certificate)

        self.delete_btn = QPushButton("🗑️ Διαγραφή")
        self.delete_btn.setProperty("role", "delete")
        self.delete_btn.setIcon(self.style().standardIcon(getattr(self.style(), 'SP_TrashIcon', None)))
        self.delete_btn.clicked.connect(self.delete_certificate)

        self.clear_btn = QPushButton("🔄 Καθαρισμός")
        self.clear_btn.setProperty("role", "clear")
        self.clear_btn.setIcon(self.style().standardIcon(getattr(self.style(), 'SP_DialogResetButton', None)))
        self.clear_btn.clicked.connect(self.clear_form)

        self.attachments_btn = QPushButton("📎 Συνημμένα")
        self.attachments_btn.setProperty("role", "attachment")
        self.attachments_btn.setIcon(self.style().standardIcon(getattr(self.style(), 'SP_FileIcon', None)))
        self.attachments_btn.clicked.connect(self.manage_attachments)

        buttons_layout.addWidget(self.add_btn)
        buttons_layout.addWidget(self.update_btn)
        buttons_layout.addWidget(self.delete_btn)
        buttons_layout.addWidget(self.clear_btn)
        buttons_layout.addWidget(self.attachments_btn)
        buttons_layout.addStretch()

        form_layout.addLayout(buttons_layout)
        main_layout.addWidget(form_group)

        # === ΟΜΑΔΑ ΑΝΑΖΗΤΗΣΗΣ & ΤΑΞΙΝΟΜΗΣΗΣ (ΜΟΝΟ ΜΙΑ ΓΡΑΜΜΗ) ===
        search_group = QGroupBox("🔍 Αναζήτηση & Ταξινόμηση")
        search_group.setStyleSheet("""
            QGroupBox {
                font-size: 12px;
                font-weight: bold;
                color: #27ae60;
            }
        """)
        search_layout = QVBoxLayout()
        search_group.setLayout(search_layout)

        # ΜΙΑ ΓΡΑΜΜΗ ΜΕ ΑΝΑΖΗΤΗΣΗ ΚΑΙ ΤΑΞΙΝΟΜΗΣΗ
        search_sort_row = QHBoxLayout()
        search_sort_row.setSpacing(15)

        # Αναζήτηση (αριστερά)
        search_label = QLabel("🔎 Αναζήτηση:")
        search_label.setStyleSheet("font-weight: bold; color: #2c3e50;")

        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Αναζήτηση με Προσωπικό Αριθμό, Όνομα, ΑΜΚΑ...")
        self.search_input.setClearButtonEnabled(True)
        self.search_input.textChanged.connect(self.on_search)
        self.search_input.setMinimumWidth(250)  # Μικρότερο πλάτος

        # Ταξινόμηση (δεξιά)
        sort_label = QLabel("📊 Ταξινόμηση:")
        sort_label.setStyleSheet("font-weight: bold; color: #2c3e50;")

        self.sort_combo = QComboBox()
        self.sort_combo.addItems([
            "📅 Ημερομηνία (Νεότερο→Παλαιότερο)",
            "📅 Ημερομηνία (Παλαιότερο→Νεότερο)",
            "📅 Έτος (Νεότερο→Παλαιότερο)",
            "📅 Έτος (Παλαιότερο→Νεότερο)",
            "🔤 Επώνυμο (Α→Ω)",
            "🔤 Επώνυμο (Ω→Α)",
            "🔢 Πρωτόκολλο (Αύξον - Τρέχον Έτος)",
            "🔢 Πρωτόκολλο (Φθίνον - Τρέχον Έτος)",
            "🔢 Πρωτόκολλο (Αύξον - Όλα τα Έτη)",
            "🔢 Πρωτόκολλο (Φθίνον - Όλα τα Έτη)"
        ])
        self.sort_combo.currentIndexChanged.connect(self.on_sort_changed)
        self.sort_combo.setMinimumWidth(200)  # Μικρότερο πλάτος

        # Προσθήκη widgets στη γραμμή
        search_sort_row.addWidget(search_label)
        search_sort_row.addWidget(self.search_input)
        search_sort_row.addStretch()  # Ελεύθερος χώρος στο κέντρο
        search_sort_row.addWidget(sort_label)
        search_sort_row.addWidget(self.sort_combo)

        search_layout.addLayout(search_sort_row)
        main_layout.addWidget(search_group)

        # === ΠΙΝΑΚΑΣ ΠΙΣΤΟΠΟΙΗΤΙΚΩΝ ===
        table_group = QGroupBox("📋 Κατάλογος Πιστοποιητικών")
        table_group.setStyleSheet("""
                QGroupBox {
                    font-size: 13px;
                    font-weight: bold;
                    color: #9b59b6;
                }
            """)
        table_layout = QVBoxLayout()
        table_group.setLayout(table_layout)

        self.certificates_table = QTreeWidget()
        
        # 🔽 ΑΛΛΑΓΗ: 14 στήλες τώρα (λόγω προσθήκης Σχολίων)
        self.certificates_table.setColumnCount(14)
        self.certificates_table.setHeaderLabels([
            "🔑 ID", "📄 Πρωτόκολλο", "👤 Επώνυμο", "👤 Όνομα", "👨 Πατρώνυμο",
            "🆔 ΑΜΚΑ", "🔢 Προσωπικός", "📇 Ταυτότητα", "🌍 Διαβατήριο",
            "📅 Ημερομηνία", "🏠 Κατοικία", "💼 Εργασία", "💬 Σχόλια", "📎 Συνημμένα"  # 🔽 ΠΡΟΣΘΗΚΗ ΣΧΟΛΙΩΝ
        ])
        self.certificates_table.setColumnHidden(0, True)
        # Ρύθμιση για πολλαπλές επιλογές
        self.certificates_table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.certificates_table.setSelectionBehavior(QAbstractItemView.SelectRows)

        # Προσθήκη εναλλασσόμενων χρωμάτων
        self.certificates_table.setAlternatingRowColors(True)

        # Αυτόματο προσαρμογή πλάτους
        self.certificates_table.header().setStretchLastSection(False)
        self.certificates_table.header().setSectionResizeMode(1, QHeaderView.ResizeToContents)  # Πρωτόκολλο
        self.certificates_table.header().setSectionResizeMode(2, QHeaderView.ResizeToContents)  # Επώνυμο
        self.certificates_table.header().setSectionResizeMode(5, QHeaderView.ResizeToContents)  # ΑΜΚΑ
        self.certificates_table.header().setSectionResizeMode(6, QHeaderView.ResizeToContents)  # Προσωπικός
        self.certificates_table.header().setSectionResizeMode(7, QHeaderView.ResizeToContents)  # Ταυτότητα
        self.certificates_table.header().setSectionResizeMode(8, QHeaderView.ResizeToContents)  # Διαβατήριο
        self.certificates_table.header().setSectionResizeMode(9, QHeaderView.ResizeToContents)  # Ημερομηνία
        self.certificates_table.header().setSectionResizeMode(12, QHeaderView.ResizeToContents)  # Σχόλια
        self.certificates_table.header().setSectionResizeMode(13, QHeaderView.Stretch)  # Συνημμένα

        self.certificates_table.itemSelectionChanged.connect(self.on_certificate_select)
        # Διπλό κλικ για άνοιγμα συνημμένων (τώρα είναι στη στήλη 13)
        self.certificates_table.itemDoubleClicked.connect(self.on_double_click_attachment)

        table_layout.addWidget(self.certificates_table)
        main_layout.addWidget(table_group)
        nav_layout = QHBoxLayout()

        self.prev_btn = QPushButton("⬅ Προηγούμενα")
        self.next_btn = QPushButton("Επόμενα ➡")

        nav_layout.addWidget(self.prev_btn)
        nav_layout.addWidget(self.next_btn)

        main_layout.addLayout(nav_layout)

        # === ΚΟΥΜΠΙΑ ΕΞΑΓΩΓΗΣ ===
        export_group = QGroupBox("📤 Εξαγωγή & Backup")
        export_group.setStyleSheet("""
                QGroupBox {
                    font-size: 12px;
                    font-weight: bold;
                    color: #e74c3c;
                }
            """)
        export_layout = QHBoxLayout()
        export_group.setLayout(export_layout)

        self.export_excel_btn = QPushButton("📊 Εξαγωγή Excel (Όλα)")
        self.export_excel_btn.setProperty("role", "export")
        self.export_excel_btn.clicked.connect(self.export_to_excel)

        self.export_selected_excel_btn = QPushButton("📊 Εξαγωγή Excel (Επιλεγμένα)")
        self.export_selected_excel_btn.setProperty("role", "export")
        self.export_selected_excel_btn.clicked.connect(self.export_selected_to_excel)

        # 🔽 ΝΕΟ ΚΟΥΜΠΙ BACKUP CSV
        self.backup_csv_btn = QPushButton("💾 Backup CSV")
        self.backup_csv_btn.setProperty("role", "export")
        self.backup_csv_btn.setStyleSheet("""
            QPushButton {
                background-color: #1abc9c;
            }
            QPushButton:hover {
                background-color: #16a085;
            }
        """)
        self.backup_csv_btn.clicked.connect(self.backup_to_csv)
        
        # 🔽 ΚΟΥΜΠΙ ΕΚΤΥΠΩΣΗΣ (ΠΡΕΠΕΙ ΝΑ ΜΕΙΝΕΙ ΛΕΙΤΟΥΡΓΙΚΟ)
        self.print_btn = QPushButton("🖨️ Εκτύπωση Φόρμας")
        self.print_btn.setProperty("role", "export")
        self.print_btn.clicked.connect(self.print_certificate_form)
        
        self.next_btn.clicked.connect(self.next_page)
        self.prev_btn.clicked.connect(self.prev_page)
        
        export_layout.addWidget(self.export_excel_btn)
        export_layout.addWidget(self.export_selected_excel_btn)
        export_layout.addWidget(self.backup_csv_btn)  # 🔽 ΠΡΟΣΘΗΚΗ
        export_layout.addWidget(self.print_btn)
        export_layout.addStretch()

        main_layout.addWidget(export_group)

    def next_page(self):
        self.current_offset += self.page_size
        self.load_certificates()

    def prev_page(self):
        if self.current_offset >= self.page_size:
            self.current_offset -= self.page_size
        self.load_certificates()

    def on_double_click_attachment(self, item, column):
        """Άνοιγμα συνημμένων με διπλό κλικ στη στήλη συνημμένων."""
        # 🔽 ΑΛΛΑΓΗ: Στήλη 13 τώρα (λόγω προσθήκης Σχολίων)
        if column == 13:  # Στήλη συνημμένων
            # Βρες τη γραμμή του κλικ
            if item.parent() is None:
                cert_item = item
            else:
                cert_item = item.parent()

            cert_id = cert_item.text(0)
            dialog = AttachmentDialog(self, cert_id)
            dialog.exec_()
            # Ανανέωση μετρητή συνημμένων
            self.update_attachments_count(cert_id)

    def print_certificate(self):
        """Εκτυπώνει το επιλεγμένο πιστοποιητικό."""
        selected = self.certificates_table.selectedItems()
        if not selected:
            QMessageBox.warning(self, "Προσοχή",
                                "Επιλέξτε ένα πιστοποιητικό για εκτύπωση.")
            return

        cert_id = selected[0].text(0)

        conn = Database.get_connection()
        if not conn:
            return

        try:
            cur = conn.cursor()
            cur.execute("""
                SELECT protocol_number, hc_first_name, hc_last_name,
                       hc_father_name, hc_amka, certificate_date,
                       residence, work_type, comments
                FROM health_certificates
                WHERE id = ? AND doctor_id = ?
            """, (cert_id, self.current_user_id))

            cert = cur.fetchone()
            if cert:
                # Δημιουργία απλού μηνύματος για εκτύπωση
                message = f"""
                ΠΙΣΤΟΠΟΙΗΤΙΚΟ ΥΓΕΙΑΣ
                =====================

                Πρωτόκολλο: {cert[0]}
                Ασθενής: {cert[2]} {cert[1]}
                Πατρώνυμο: {cert[3] if cert[3] else '-'}
                ΑΜΚΑ: {cert[4] if cert[4] else '-'}
                Ημερομηνία: {cert[5].strftime('%d/%m/%Y') if cert[5] else '-'}
                Κατοικία: {cert[6]}
                Είδος Εργασίας: {cert[7]}

                Σχόλια:
                {cert[8] if cert[8] else '-'}

                ---------------------
                Υπογραφή Ιατρού
                """

                # Άνοιγμα dialog για προβολή/εκτύπωση
                from PyQt5.QtWidgets import QTextEdit, QDialog, QVBoxLayout, QPushButton, QPrintDialog
                from PyQt5.QtPrintSupport import QPrinter

                dialog = QDialog(self)
                dialog.setWindowTitle(f"Εκτύπωση Πιστοποιητικού: {cert[0]}")
                dialog.resize(500, 600)

                layout = QVBoxLayout()

                text_edit = QTextEdit()
                text_edit.setPlainText(message)
                text_edit.setReadOnly(True)

                print_btn = QPushButton("🖨️ Εκτύπωση")
                print_btn.clicked.connect(lambda: self._print_text(text_edit.toPlainText()))

                close_btn = QPushButton("Κλείσιμο")
                close_btn.clicked.connect(dialog.accept)

                layout.addWidget(text_edit)
                layout.addWidget(print_btn)
                layout.addWidget(close_btn)

                dialog.setLayout(layout)
                dialog.exec_()
            else:
                QMessageBox.warning(self, "Προσοχή", "Το πιστοποιητικό δεν βρέθηκε.")

        except Exception as e:
            QMessageBox.critical(self, "Σφάλμα", f"Σφάλμα εκτύπωσης: {e}")

    def print_certificate(self):
        """Εκτυπώνει το επιλεγμένο πιστοποιητικό."""
        selected = self.certificates_table.selectedItems()
        if not selected:
            QMessageBox.warning(self, "Προσοχή",
                                "Επιλέξτε ένα πιστοποιητικό για εκτύπωση.")
            return

        cert_id = selected[0].text(0)

        conn = Database.get_connection()
        if not conn:
            return

        try:
            cur = conn.cursor()
            cur.execute("""
                SELECT protocol_number, hc_first_name, hc_last_name,
                       hc_father_name, hc_amka, certificate_date,
                       residence, work_type, comments
                FROM health_certificates
                WHERE id = ? AND doctor_id = ?
            """, (cert_id, self.current_user_id))

            cert = cur.fetchone()
            if cert:
                # Δημιουργία απλού μηνύματος για εκτύπωση
                message = f"""
                ΠΙΣΤΟΠΟΙΗΤΙΚΟ ΥΓΕΙΑΣ
                =====================

                Πρωτόκολλο: {cert[0]}
                Ασθενής: {cert[2]} {cert[1]}
                Πατρώνυμο: {cert[3] if cert[3] else '-'}
                ΑΜΚΑ: {cert[4] if cert[4] else '-'}
                Ημερομηνία: {cert[5].strftime('%d/%m/%Y') if cert[5] else '-'}
                Κατοικία: {cert[6]}
                Είδος Εργασίας: {cert[7]}

                Σχόλια:
                {cert[8] if cert[8] else '-'}

                ---------------------
                Υπογραφή Ιατρού
                """

                # Άνοιγμα dialog για προβολή/εκτύπωση
                from PyQt5.QtWidgets import QTextEdit, QDialog, QVBoxLayout, QPushButton, QPrintDialog
                from PyQt5.QtPrintSupport import QPrinter

                dialog = QDialog(self)
                dialog.setWindowTitle(f"Εκτύπωση Πιστοποιητικού: {cert[0]}")
                dialog.resize(500, 600)

                layout = QVBoxLayout()

                text_edit = QTextEdit()
                text_edit.setPlainText(message)
                text_edit.setReadOnly(True)

                print_btn = QPushButton("🖨️ Εκτύπωση")
                print_btn.clicked.connect(lambda: self._print_text(text_edit.toPlainText()))

                close_btn = QPushButton("Κλείσιμο")
                close_btn.clicked.connect(dialog.accept)

                layout.addWidget(text_edit)
                layout.addWidget(print_btn)
                layout.addWidget(close_btn)

                dialog.setLayout(layout)
                dialog.exec_()
            else:
                QMessageBox.warning(self, "Προσοχή", "Το πιστοποιητικό δεν βρέθηκε.")

        except Exception as e:
            QMessageBox.critical(self, "Σφάλμα", f"Σφάλμα εκτύπωσης: {e}")

    def setup_form_fields(self, layout):
        """Δημιουργεί τα πεδία της φόρμας με icons."""

        # Γραμμή 1 - Πρωτόκολλο (αυτόματο) και Επώνυμο (δίπλα)
        row1 = QHBoxLayout()
        row1.setSpacing(10)

        proto_label = QLabel("📄 Πρωτόκολλο:")
        proto_label.setStyleSheet("font-weight: bold;")
        self.protocol_input = QLineEdit()
        self.protocol_input.setReadOnly(True)
        self.protocol_input.setStyleSheet("background-color: #f0f0f0;")
        self.protocol_input.setMaximumWidth(200)

        self.refresh_proto_btn = QPushButton("🔄 Νέο")
        self.refresh_proto_btn.setProperty("role", "update")
        self.refresh_proto_btn.setMaximumWidth(80)
        self.refresh_proto_btn.clicked.connect(self.refresh_protocol)
        
        self.change_date_btn = QPushButton("📅 Αλλαγή Ημερομηνίας")
        self.change_date_btn.setProperty("role", "update")
        self.change_date_btn.setMaximumWidth(120)
        self.change_date_btn.clicked.connect(self.change_protocol_date)

        lastname_label = QLabel("👤 Επώνυμο:")
        lastname_label.setStyleSheet("font-weight: bold;")
        self.lastname_input = QLineEdit()
        self.lastname_input.setPlaceholderText("Επώνυμο ασθενούς")

        row1.addWidget(proto_label)
        row1.addWidget(self.protocol_input)
        row1.addWidget(self.refresh_proto_btn)
        row1.addWidget(self.change_date_btn)  # 🔽 ΠΡΟΣΘΗΚΗ
        row1.addSpacing(20)
        row1.addWidget(lastname_label)
        row1.addWidget(self.lastname_input)
        row1.addStretch()

        layout.addLayout(row1)

        # Γραμμή 2 - Όνομα και Πατρώνυμο
        row2 = QHBoxLayout()
        row2.setSpacing(10)

        firstname_label = QLabel("👤 Όνομα:")
        firstname_label.setStyleSheet("font-weight: bold;")
        self.firstname_input = QLineEdit()
        self.firstname_input.setPlaceholderText("Όνομα ασθενούς")

        fathername_label = QLabel("👨 Πατρώνυμο:")
        fathername_label.setStyleSheet("font-weight: bold;")
        self.fathername_input = QLineEdit()
        self.fathername_input.setPlaceholderText("Όνομα πατρός")

        row2.addWidget(firstname_label)
        row2.addWidget(self.firstname_input)
        row2.addWidget(fathername_label)
        row2.addWidget(self.fathername_input)
        row2.addStretch()

        layout.addLayout(row2)

        # Γραμμή 3 - ΑΜΚΑ και Προσωπικός Αριθμός
        row3 = QHBoxLayout()
        row3.setSpacing(10)

        amka_label = QLabel("🆔 ΑΜΚΑ:")
        amka_label.setStyleSheet("font-weight: bold;")
        self.amka_input = QLineEdit()
        self.amka_input.setPlaceholderText("11 ψηφία (ή κενό)")
        self.amka_input.setMaxLength(11)  # ΑΚΡΙΒΩΣ 11 χαρακτήρες
        self.amka_input.setMaximumWidth(150)
        # Validator για ΑΚΡΙΒΩΣ 11 ψηφία
        amka_validator = QRegExpValidator(QRegExp(r"\d{0,11}"))
        self.amka_input.setValidator(amka_validator)

        personal_label = QLabel("🔢 Προσωπικός:")
        personal_label.setStyleSheet("font-weight: bold;")
        self.personal_input = QLineEdit()
        self.personal_input.setPlaceholderText("12 χαρακτήρες (γράμματα & ψηφία)")
        self.personal_input.setMaxLength(12)  # ΑΚΡΙΒΩΣ 12 χαρακτήρες
        self.personal_input.setMaximumWidth(150)
        # 🔽 ΔΙΟΡΘΩΣΗ: Validator για ελληνικά, λατινικά γράμματα και ψηφία
        personal_validator = QRegExpValidator(QRegExp(r"[Α-Ωα-ωA-Za-z0-9]{0,12}"))
        self.personal_input.setValidator(personal_validator)

        row3.addWidget(amka_label)
        row3.addWidget(self.amka_input)
        row3.addSpacing(20)
        row3.addWidget(personal_label)
        row3.addWidget(self.personal_input)
        row3.addStretch()

        layout.addLayout(row3)

        # Γραμμή 4 - Ταυτότητα και Διαβατήριο (πιο στενά)
        row4 = QHBoxLayout()
        row4.setSpacing(10)

        id_label = QLabel("📇 Ταυτότητα:")
        id_label.setStyleSheet("font-weight: bold;")
        self.id_input = QLineEdit()
        self.id_input.setPlaceholderText("Αρ. Ταυτότητας")
        self.id_input.setMaximumWidth(120)

        passport_label = QLabel("🌍 Διαβατήριο:")
        passport_label.setStyleSheet("font-weight: bold;")
        self.passport_input = QLineEdit()
        self.passport_input.setPlaceholderText("Αρ. Διαβατηρίου")
        self.passport_input.setMaximumWidth(120)

        row4.addWidget(id_label)
        row4.addWidget(self.id_input)
        row4.addSpacing(20)
        row4.addWidget(passport_label)
        row4.addWidget(self.passport_input)
        row4.addStretch()

        layout.addLayout(row4)

        # Γραμμή 5 - Ημερομηνία, Κατοικία, Εργασία (3 ΣΤΗΛΕΣ)
        row5 = QHBoxLayout()
        row5.setSpacing(10)

        # Στήλη 1: Ημερομηνία
        date_label = QLabel("📅 Ημερομηνία:")
        date_label.setStyleSheet("font-weight: bold;")
        self.date_input = QDateEdit()
        self.date_input.setDate(QDate.currentDate())
        self.date_input.setCalendarPopup(True)
        self.date_input.setDisplayFormat("dd/MM/yyyy")
        self.date_input.setMaximumWidth(120)

        # Στήλη 2: Κατοικία
        residence_label = QLabel("🏠 Κατοικία:")
        residence_label.setStyleSheet("font-weight: bold;")
        self.residence_input = QLineEdit()
        self.residence_input.setPlaceholderText("Διεύθυνση")
        self.residence_input.setMaximumWidth(180)

        # Στήλη 3: Εργασία
        work_label = QLabel("💼 Εργασία:")
        work_label.setStyleSheet("font-weight: bold;")
        self.work_input = QLineEdit()
        self.work_input.setPlaceholderText("Είδος εργασίας")
        self.work_input.setMaximumWidth(180)

        # Προσθήκη και στις 3 στήλες
        row5.addWidget(date_label)
        row5.addWidget(self.date_input)
        row5.addSpacing(20)
        row5.addWidget(residence_label)
        row5.addWidget(self.residence_input)
        row5.addSpacing(20)
        row5.addWidget(work_label)
        row5.addWidget(self.work_input)
        row5.addStretch()

        layout.addLayout(row5)

        # Σχόλια (λίγο μικρότερο)
        comments_label = QLabel("💬 Σχόλια:")
        comments_label.setStyleSheet("font-weight: bold;")
        layout.addWidget(comments_label)

        self.comments_input = QTextEdit()
        self.comments_input.setPlaceholderText("Προαιρετικά σχόλια...")
        self.comments_input.setMaximumHeight(60)  # Μικρότερο από 80
        layout.addWidget(self.comments_input)

    def refresh_protocol(self):
        """Δημιουργεί νέο πρωτόκολλο με το μεγαλύτερο αριθμό για το τρέχον έτος."""
        try:
            from datetime import datetime
            current_year = datetime.now().year

            conn = Database.get_connection()
            if not conn:
                return

            try:
                cur = conn.cursor()

                # Βρες τον τελευταίο αριθμό πρωτοκόλλου για το τρέχον έτος
                cur.execute("""
                    SELECT MAX(proto_seq) as max_seq
                    FROM health_certificates
                    WHERE doctor_id = ? AND proto_year = ?
                """, (self.current_user_id, current_year))

                result = cur.fetchone()

                if result and result[0] is not None:
                    next_seq = int(result[0]) + 1
                else:
                    next_seq = 1

                # Μορφοποίηση: ΝΝΝΝ/DD-MM-YYYY
                today = datetime.now()
                protocol_number = f"{next_seq:04d}/{today.day:02d}-{today.month:02d}-{today.year}"

                self.protocol_input.setText(protocol_number)

            except Exception as e:
                print(f"Σφάλμα ανανέωσης πρωτοκόλλου: {e}")
                import traceback
                traceback.print_exc()

        except Exception as e:
            print(f"Σφάλμα ανανέωσης πρωτοκόλλου: {e}")
            import traceback
            traceback.print_exc()

    from PyQt5.QtCore import QPropertyAnimation, QEasingCurve

    def animate_button(self, button):
        """Προσθέτει animation στα κουμπιά."""
        animation = QPropertyAnimation(button, b"geometry")
        animation.setDuration(200)
        animation.setEasingCurve(QEasingCurve.OutBack)

        original_geom = button.geometry()
        animation.setStartValue(original_geom)
        animation.setEndValue(original_geom)

        animation.start()


    def extract_year_from_date(self, date_str):
        """Εξάγει το έτος από μια ημερομηνία string (για μελλοντική χρήση)."""
        if not date_str:
            return 0
        try:
            if isinstance(date_str, str):
                # Format: YYYY-MM-DD
                if '-' in date_str:
                    return int(date_str.split('-')[0])
                # Format: DD/MM/YYYY
                elif '/' in date_str:
                    parts = date_str.split('/')
                    if len(parts) == 3:
                        return int(parts[2])
                # Απλός αριθμός
                elif date_str.isdigit() and len(date_str) >= 4:
                    return int(date_str[:4])
            return 0
        except:
            return 0

    def load_certificates(self):
        """Φορτώνει τα πιστοποιητικά από τη βάση."""
        self.certificates_table.clear()
        self.certificates_table.setUpdatesEnabled(False)
        self.certificates_table.blockSignals(True)
        
        # 🔽 ΕΝΗΜΕΡΩΣΗ: 14 στήλες με Σχόλια
        self.certificates_table.setColumnCount(14)
        self.certificates_table.setHeaderLabels([
            "🔑 ID", "📄 Πρωτόκολλο", "👤 Επώνυμο", "👤 Όνομα", "👨 Πατρώνυμο",
            "🆔 ΑΜΚΑ", "🔢 Προσωπικός", "📇 Ταυτότητα", "🌍 Διαβατήριο",
            "📅 Ημερομηνία", "🏠 Κατοικία", "💼 Εργασία", "💬 Σχόλια", "📎 Συνημμένα"
        ])
        self.certificates_table.setColumnHidden(0, True)

        conn = Database.get_connection()
        if not conn:
            return

        try:
            cur = conn.cursor()

            # Ταξινόμηση βάσει επιλογής
            sort_by = self.sort_combo.currentText()

            # Εισαγωγή datetime για χρήση σε τρέχον έτος
            from datetime import datetime
            current_year = datetime.now().year

            if sort_by == "📅 Ημερομηνία (Νεότερο→Παλαιότερο)":
                order_by = "certificate_date DESC, proto_year DESC, proto_seq DESC"
            elif sort_by == "📅 Ημερομηνία (Παλαιότερο→Νεότερο)":
                order_by = "certificate_date ASC, proto_year ASC, proto_seq ASC"
            elif sort_by == "📅 Έτος (Νεότερο→Παλαιότερο)":
                order_by = "proto_year DESC, proto_seq DESC"
            elif sort_by == "📅 Έτος (Παλαιότερο→Νεότερο)":
                order_by = "proto_year ASC, proto_seq ASC"
            elif sort_by == "🔤 Επώνυμο (Α→Ω)":
                order_by = "hc_last_name ASC, hc_first_name ASC"
            elif sort_by == "🔤 Επώνυμο (Ω→Α)":
                order_by = "hc_last_name DESC, hc_first_name DESC"
            elif sort_by == "🔢 Πρωτόκολλο (Αύξον - Τρέχον Έτος)":
                # Χρησιμοποιούμε παράμετρο για το current_year
                order_by = f"CASE WHEN proto_year = ? THEN 0 ELSE 1 END, proto_year DESC, proto_seq ASC"
            elif sort_by == "🔢 Πρωτόκολλο (Φθίνον - Τρέχον Έτος)":
                order_by = f"CASE WHEN proto_year = ? THEN 0 ELSE 1 END, proto_year DESC, proto_seq DESC"
            elif sort_by == "🔢 Πρωτόκολλο (Αύξον - Όλα τα Έτη)":
                order_by = "proto_year ASC, proto_seq ASC"
            elif sort_by == "🔢 Πρωτόκολλο (Φθίνον - Όλα τα Έτη)":
                order_by = "proto_year DESC, proto_seq DESC"
            else:
                order_by = "certificate_date DESC, proto_year DESC, proto_seq DESC"

            # 🔽 ΕΝΗΜΕΡΩΣΗ QUERY: Προσθήκη comments και χρήση σωστής order_by
            query = f"""
                SELECT id, protocol_number, hc_last_name, hc_first_name,
                       hc_father_name, hc_amka, hc_personal_number,
                       id_number, passport_number, certificate_date,
                       residence, work_type, comments
                FROM health_certificates
                WHERE doctor_id = ?
                ORDER BY {order_by}
                LIMIT ? OFFSET ?
            """

            # Εκτέλεση με διαφορετικές παραμέτρους ανάλογα με την ταξινόμηση
            if "Τρέχον Έτος" in sort_by:
                cur.execute(query, (self.current_user_id, current_year, self.page_size, self.current_offset))
            else:
                cur.execute(query, (self.current_user_id, self.page_size, self.current_offset))

            certificates = cur.fetchall()

            for cert in certificates:
                item = QTreeWidgetItem()
                self.certificates_table.addTopLevelItem(item)

                # 🔽 ΕΝΗΜΕΡΩΜΕΝΕΣ ΘΕΣΕΙΣ ΛΟΓΩ ΠΡΟΣΘΗΚΗΣ ΣΧΟΛΙΩΝ
                item.setText(0, str(cert[0]))  # ID
                item.setText(1, cert[1] if cert[1] else "")  # Πρωτόκολλο
                item.setText(2, cert[2] if cert[2] else "")  # Επώνυμο
                item.setText(3, cert[3] if cert[3] else "")  # Όνομα
                item.setText(4, cert[4] if cert[4] else "")  # Πατρώνυμο
                item.setText(5, cert[5] if cert[5] else "")  # ΑΜΚΑ
                item.setText(6, cert[6] if cert[6] else "")  # Προσωπικός
                item.setText(7, cert[7] if cert[7] else "")  # Ταυτότητα
                item.setText(8, cert[8] if cert[8] else "")  # Διαβατήριο

                # Χειρισμός ημερομηνίας (τώρα είναι στη θέση 9)
                date_str = cert[9] if cert[9] else ""
                if date_str:
                    try:
                        if isinstance(date_str, str):
                            try:
                                date_obj = datetime.strptime(date_str, "%Y-%m-%d")
                            except:
                                date_obj = datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
                            date_str = date_obj.strftime("%d/%m/%Y")
                        else:
                            date_str = date_str.strftime("%d/%m/%Y")
                    except Exception:
                        date_str = str(date_str)[:10]
                item.setText(9, date_str)  # Ημερομηνία

                item.setText(10, cert[10] if cert[10] else "")  # Κατοικία
                item.setText(11, cert[11] if cert[11] else "")  # Εργασία

                # 🔽 ΠΡΟΣΘΗΚΗ: Σχόλια (στη στήλη 12)
                comments = cert[12] if cert[12] else ""
                # Κόψε τα σχόλια αν είναι πολύ μεγάλα για καλύτερη εμφάνιση
                if len(comments) > 50:
                    item.setText(12, comments[:47] + "...")
                    item.setToolTip(12, comments)  # Πλήρες κείμενο ως tooltip
                else:
                    item.setText(12, comments)

                item.setText(13, "")  # Συνημμένα (τώρα στη στήλη 13)

            # Ανανέωση συνημμένων
            for i in range(self.certificates_table.topLevelItemCount()):
                item = self.certificates_table.topLevelItem(i)
                cert_id = item.text(0)
                self.update_attachments_count(cert_id)
                self.certificates_table.blockSignals(False)
                self.certificates_table.setUpdatesEnabled(True)
                self.certificates_table.viewport().update()

        except Exception as e:
            QMessageBox.critical(self, "Σφάλμα", f"Σφάλμα φόρτωσης: {e}")
            import traceback
            traceback.print_exc()

    def update_attachments_count(self, certificate_id):
        """Ενημερώνει τη στήλη συνημμένων."""
        conn = Database.get_connection()
        if not conn:
            return

        try:
            cur = conn.cursor()

            cur.execute("""
                SELECT COUNT(*) FROM health_certificate_documents
                WHERE certificate_id = ? AND doctor_id = ?
            """, (certificate_id, self.current_user_id))

            count = cur.fetchone()[0]

            # Βρες το item
            for i in range(self.certificates_table.topLevelItemCount()):
                item = self.certificates_table.topLevelItem(i)
                if item.text(0) == certificate_id:
                    if count > 0:
                        # Πάρε το όνομα του πρώτου αρχείου
                        cur.execute("""
                            SELECT file_name FROM health_certificate_documents
                            WHERE certificate_id = ? AND doctor_id = ?
                            ORDER BY upload_date DESC LIMIT 1
                        """, (certificate_id, self.current_user_id))

                        file_result = cur.fetchone()
                        if file_result:
                            file_name = file_result[0]
                            if len(file_name) > 15:
                                display = file_name[:12] + "..."
                            else:
                                display = file_name

                            if count > 1:
                                item.setText(13, f"{display} (+{count - 1})")  # 🔽 ΣΤΗΛΗ 13 ΤΩΡΑ
                            else:
                                item.setText(13, display)  # 🔽 ΣΤΗΛΗ 13 ΤΩΡΑ
                    else:
                        item.setText(13, "")  # 🔽 ΣΤΗΛΗ 13 ΤΩΡΑ
                    break

        except Exception as e:
            print(f"Σφάλμα ενημέρωσης συνημμένων: {e}")
            import traceback
            traceback.print_exc()

    def add_certificate(self):
        """Προσθέτει νέο πιστοποιητικό."""
        # Έλεγχος υποχρεωτικών πεδίων
        required_fields = [
            (self.protocol_input, "Πρωτόκολλο"),
            (self.lastname_input, "Επώνυμο"),
            (self.firstname_input, "Όνομα"),
            (self.residence_input, "Κατοικία"),
            (self.work_input, "Είδος εργασίας")
        ]

        for field, name in required_fields:
            if not field.text().strip():
                QMessageBox.warning(self, "Σφάλμα",
                                    f"Το πεδίο '{name}' είναι υποχρεωτικό.")
                field.setFocus()
                return

        # Έλεγχος ΑΜΚΑ - ΑΚΡΙΒΩΣ 11 ψηφία αν δεν είναι κενό
        amka = self.amka_input.text().strip()
        if amka:  # Μόνο αν δεν είναι κενό
            if not re.fullmatch(r"\d{11}", amka):
                QMessageBox.warning(self, "Σφάλμα",
                                    "Ο ΑΜΚΑ πρέπει να έχει ΑΚΡΙΒΩΣ 11 ψηφία.\n"
                                    "Ή αφήστε το πεδίο κενό αν δεν υπάρχει ΑΜΚΑ.")
                self.amka_input.setFocus()
                return

            # Έλεγχος αν ο ΑΜΚΑ υπάρχει ήδη
            existing_cert = self.check_existing_amka(amka)
            if existing_cert:
                reply = self.show_amka_conflict_dialog(existing_cert, amka)
                if reply == "cancel":
                    return
                elif reply == "replace":
                    if not self.delete_existing_certificate(existing_cert[0]):
                        return

        # Έλεγχος Προσωπικού Αριθμού - ΑΚΡΙΒΩΣ 12 χαρακτήρες αν δεν είναι κενό
        personal_number = self.personal_input.text().strip()
        if personal_number:  # Μόνο αν δεν είναι κενό
            # 🔽 ΔΙΟΡΘΩΣΗ: ΑΚΡΙΒΩΣ 12 χαρακτήρες (ελληνικά/λατινικά γράμματα & ψηφία)
            if not re.fullmatch(r"[Α-Ωα-ωA-Za-z0-9]{12}", personal_number):
                QMessageBox.warning(self, "Σφάλμα",
                                    "Ο Προσωπικός Αριθμός πρέπει να έχει ΑΚΡΙΒΩΣ 12 χαρακτήρες.\n"
                                    "Επιτρέπονται: ελληνικά/λατινικά γράμματα και ψηφία.\n"
                                    "Ή αφήστε το πεδίο κενό αν δεν υπάρχει Προσωπικός Αριθμός.")
                self.personal_input.setFocus()
                return

            # Έλεγχος αν ο Προσωπικός Αριθμός υπάρχει ήδη
            existing_personal = self.check_existing_personal_number(personal_number)
            if existing_personal:
                reply = self.show_personal_conflict_dialog(existing_personal, personal_number)
                if reply == "cancel":
                    return
                elif reply == "replace":
                    if not self.delete_existing_certificate(existing_personal[0]):
                        return

        conn = Database.get_connection()
        if not conn:
            return

        try:
            cur = conn.cursor()

            # Parse πρωτόκολλο
            proto_match = re.match(r"(\d+)/(\d+)-(\d+)-(\d+)",
                                   self.protocol_input.text().strip())
            if not proto_match:
                QMessageBox.warning(self, "Σφάλμα",
                                    "Λάθος μορφή πρωτοκόλλου. Χρήση: ΝΝΝΝ/DD-MM-YYYY")
                return

            seq = int(proto_match.group(1))
            day = int(proto_match.group(2))
            month = int(proto_match.group(3))
            year = int(proto_match.group(4))

            # Δημιουργία μοναδικού ID
            cert_id = str(uuid.uuid4())

            # Εισαγωγή στη βάση
            cur.execute("""
                INSERT INTO health_certificates (
                    id, protocol_number, proto_year, proto_seq,
                    hc_first_name, hc_last_name, hc_father_name,
                    hc_amka, hc_personal_number, id_number,
                    passport_number, certificate_date, residence,
                    work_type, comments, doctor_id
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                cert_id,
                self.protocol_input.text().strip(),
                year, seq,
                self.firstname_input.text().strip(),
                self.lastname_input.text().strip(),
                self.fathername_input.text().strip() or None,
                amka or None,
                personal_number or None,
                self.id_input.text().strip() or None,
                self.passport_input.text().strip() or None,
                self.date_input.date().toPyDate().isoformat(),
                self.residence_input.text().strip(),
                self.work_input.text().strip(),
                self.comments_input.toPlainText().strip() or None,
                self.current_user_id
            ))

            conn.commit()
            QMessageBox.information(self, "Επιτυχία",
                                    "Το πιστοποιητικό προστέθηκε.")

            # Ανανέωση
            self.clear_form()
            self.load_certificates()

        except sqlite3.IntegrityError as e:
            conn.rollback()
            error_msg = str(e)

            if "UNIQUE constraint failed: health_certificates.hc_amka" in error_msg:
                QMessageBox.critical(self, "Σφάλμα",
                                     f"Ο ΑΜΚΑ '{amka}' υπάρχει ήδη στη βάση.\n"
                                     f"Παρακαλώ χρησιμοποιήστε διαφορετικό ΑΜΚΑ.")
            elif "UNIQUE constraint failed: health_certificates.hc_personal_number" in error_msg:
                QMessageBox.critical(self, "Σφάλμα",
                                     f"Ο Προσωπικός Αριθμός '{personal_number}' υπάρχει ήδη στη βάση.\n"
                                     f"Παρακαλώ χρησιμοποιήστε διαφορετικό Προσωπικό Αριθμό.")
            else:
                QMessageBox.critical(self, "Σφάλμα", f"Σφάλμα βάσης δεδομένων: {e}")
        except Exception as e:
            conn.rollback()
            QMessageBox.critical(self, "Σφάλμα", f"Σφάλμα προσθήκης: {e}")
            import traceback
            traceback.print_exc()

    def check_existing_personal_number(self, personal_number):
        """Ελέγχει αν ο Προσωπικός Αριθμός υπάρχει ήδη στη βάση."""
        conn = Database.get_connection()
        if not conn:
            return None

        try:
            cur = conn.cursor()
            cur.execute("""
                SELECT id, protocol_number, hc_first_name, hc_last_name, 
                       hc_personal_number, certificate_date, doctor_id
                FROM health_certificates 
                WHERE hc_personal_number = ? AND doctor_id = ?
            """, (personal_number, self.current_user_id))

            result = cur.fetchone()
            return result  # Επιστρέφει tuple με τα στοιχεία ή None

        except Exception as e:
            print(f"Σφάλμα ελέγχου Προσωπικού Αριθμού: {e}")
            return None

    def show_personal_conflict_dialog(self, existing_data, personal_number):
        """
        Εμφανίζει διάλογο όταν υπάρχει διπλότυπος Προσωπικός Αριθμός.
        Επιστρέφει: "replace", "ignore", ή "cancel"
        """
        from PyQt5.QtWidgets import QDialog, QVBoxLayout, QLabel, QHBoxLayout, QPushButton

        dialog = QDialog(self)
        dialog.setWindowTitle("⚠️ Υπάρχον Προσωπικός Αριθμός")
        dialog.resize(450, 200)

        layout = QVBoxLayout()

        # Μήνυμα
        message = QLabel(
            f"<b>Ο Προσωπικός Αριθμός '{personal_number}' υπάρχει ήδη στη βάση:</b><br><br>"
            f"<b>Πρωτόκολλο:</b> {existing_data[1]}<br>"
            f"<b>Ασθενής:</b> {existing_data[2]} {existing_data[3]}<br>"
            f"<b>Προσ. Αριθμός:</b> {existing_data[4]}<br>"
            f"<b>Ημερομηνία:</b> {existing_data[5] if existing_data[5] else 'Άγνωστη'}<br><br>"
            f"<b>Τι θέλετε να κάνετε;</b>"
        )
        message.setWordWrap(True)

        layout.addWidget(message)

        # Κουμπιά
        buttons_layout = QHBoxLayout()

        replace_btn = QPushButton("🔄 Αντικατάσταση")
        replace_btn.setProperty("role", "delete")
        replace_btn.clicked.connect(lambda: dialog.done(1))

        ignore_btn = QPushButton("➕ Προσθήκη (Παράβλεψη)")
        ignore_btn.setProperty("role", "add")
        ignore_btn.clicked.connect(lambda: dialog.done(2))

        cancel_btn = QPushButton("❌ Ακύρωση")
        cancel_btn.clicked.connect(lambda: dialog.done(0))

        buttons_layout.addWidget(replace_btn)
        buttons_layout.addWidget(ignore_btn)
        buttons_layout.addWidget(cancel_btn)

        layout.addLayout(buttons_layout)
        dialog.setLayout(layout)

        # Εκτέλεση διαλόγου και επιστροφή αποτελέσματος
        result = dialog.exec_()

        if result == 1:
            return "replace"
        elif result == 2:
            return "ignore"
        else:
            return "cancel"

    def check_existing_amka(self, amka):
        """Ελέγχει αν ο ΑΜΚΑ υπάρχει ήδη στη βάση."""
        conn = Database.get_connection()
        if not conn:
            return None

        try:
            cur = conn.cursor()
            cur.execute("""
                SELECT id, protocol_number, hc_first_name, hc_last_name, 
                       certificate_date, doctor_id
                FROM health_certificates 
                WHERE hc_amka = ? AND doctor_id = ?
            """, (amka, self.current_user_id))

            result = cur.fetchone()
            return result  # Επιστρέφει tuple με τα στοιχεία ή None

        except Exception as e:
            print(f"Σφάλμα ελέγχου ΑΜΚΑ: {e}")
            return None

    def show_amka_conflict_dialog(self, existing_data, new_amka):
        """
        Εμφανίζει διάλογο όταν υπάρχει διπλότυπος ΑΜΚΑ.
        Επιστρέφει: "replace", "ignore", ή "cancel"
        """
        from PyQt5.QtWidgets import QDialog, QVBoxLayout, QLabel, QHBoxLayout, QPushButton

        dialog = QDialog(self)
        dialog.setWindowTitle("⚠️ Υπάρχον ΑΜΚΑ")
        dialog.resize(450, 200)

        layout = QVBoxLayout()

        # Μήνυμα
        message = QLabel(
            f"<b>Ο ΑΜΚΑ '{new_amka}' υπάρχει ήδη στη βάση:</b><br><br>"
            f"<b>Πρωτόκολλο:</b> {existing_data[1]}<br>"
            f"<b>Ασθενής:</b> {existing_data[2]} {existing_data[3]}<br>"
            f"<b>Ημερομηνία:</b> {existing_data[4] if existing_data[4] else 'Άγνωστη'}<br><br>"
            f"<b>Τι θέλετε να κάνετε;</b>"
        )
        message.setWordWrap(True)

        layout.addWidget(message)

        # Κουμπιά
        buttons_layout = QHBoxLayout()

        # Αντικατάσταση (διαγραφή παλιού, εισαγωγή νέου)
        replace_btn = QPushButton("🔄 Αντικατάσταση")
        replace_btn.setProperty("role", "delete")
        replace_btn.clicked.connect(lambda: dialog.done(1))  # Κωδικός 1 για replace

        # Προσθήκη (παραβίαση UNIQUE constraint - απενεργοποιείται)
        ignore_btn = QPushButton("➕ Προσθήκη (Παράβλεψη)")
        ignore_btn.setProperty("role", "add")
        ignore_btn.clicked.connect(lambda: dialog.done(2))  # Κωδικός 2 για ignore

        # Ακύρωση
        cancel_btn = QPushButton("❌ Ακύρωση")
        cancel_btn.clicked.connect(lambda: dialog.done(0))  # Κωδικός 0 για cancel

        buttons_layout.addWidget(replace_btn)
        buttons_layout.addWidget(ignore_btn)
        buttons_layout.addWidget(cancel_btn)

        layout.addLayout(buttons_layout)
        dialog.setLayout(layout)

        # Εκτέλεση διαλόγου και επιστροφή αποτελέσματος
        result = dialog.exec_()

        if result == 1:
            return "replace"
        elif result == 2:
            return "ignore"
        else:
            return "cancel"

    def delete_existing_certificate(self, cert_id):
        """Διαγράφει υπάρχον πιστοποιητικό πριν την αντικατάσταση."""
        reply = QMessageBox.question(
            self, "Επιβεβαίωση Διαγραφής",
            "Θέλετε να διαγραφεί το υπάρχον πιστοποιητικό πριν από την προσθήκη του νέου;\n"
            "Αυτό θα διαγράψει και όλα τα συνημμένα αρχεία.",
            QMessageBox.Yes | QMessageBox.No
        )

        if reply != QMessageBox.Yes:
            return False

        conn = Database.get_connection()
        if not conn:
            return False

        try:
            cur = conn.cursor()

            # Διαγραφή συνημμένων πρώτα
            cur.execute("""
                DELETE FROM health_certificate_documents
                WHERE certificate_id = ? AND doctor_id = ?
            """, (cert_id, self.current_user_id))

            # Διαγραφή πιστοποιητικού
            cur.execute("""
                DELETE FROM health_certificates
                WHERE id = ? AND doctor_id = ?
            """, (cert_id, self.current_user_id))

            conn.commit()
            return True

        except Exception as e:
            conn.rollback()
            QMessageBox.critical(self, "Σφάλμα", f"Σφάλμα διαγραφής: {e}")
            return False

    def update_certificate(self):
        """Ενημερώνει υπάρχον πιστοποιητικό."""
        selected = self.certificates_table.selectedItems()
        if not selected:
            QMessageBox.warning(self, "Προσοχή",
                                "Επιλέξτε ένα πιστοποιητικό για ενημέρωση.")
            return

        cert_id = selected[0].text(0)

        # Έλεγχος υποχρεωτικών πεδίων
        required_fields = [
            (self.protocol_input, "Πρωτόκολλο"),
            (self.lastname_input, "Επώνυμο"),
            (self.firstname_input, "Όνομα"),
            (self.residence_input, "Κατοικία"),
            (self.work_input, "Είδος εργασίας")
        ]

        for field, name in required_fields:
            if not field.text().strip():
                QMessageBox.warning(self, "Σφάλμα",
                                    f"Το πεδίο '{name}' είναι υποχρεωτικό.")
                field.setFocus()
                return

        # Έλεγχος ΑΜΚΑ - ΑΚΡΙΒΩΣ 11 ψηφία αν δεν είναι κενό
        amka = self.amka_input.text().strip()
        if amka:  # Μόνο αν δεν είναι κενό
            if not re.fullmatch(r"\d{11}", amka):
                QMessageBox.warning(self, "Σφάλμα",
                                    "Ο ΑΜΚΑ πρέπει να έχει ΑΚΡΙΒΩΣ 11 ψηφία.\n"
                                    "Ή αφήστε το πεδίο κενό αν δεν υπάρχει ΑΜΚΑ.")
                self.amka_input.setFocus()
                return

        # Έλεγχος Προσωπικού Αριθμού - ΑΚΡΙΒΩΣ 12 χαρακτήρες αν δεν είναι κενό
        personal_number = self.personal_input.text().strip()
        if personal_number:  # Μόνο αν δεν είναι κενό
            # 🔽 ΔΙΟΡΘΩΣΗ: ΑΚΡΙΒΩΣ 12 χαρακτήρες (ελληνικά/λατινικά γράμματα & ψηφία)
            if not re.fullmatch(r"[Α-Ωα-ωA-Za-z0-9]{12}", personal_number):
                QMessageBox.warning(self, "Σφάλμα",
                                    "Ο Προσωπικός Αριθμός πρέπει να έχει ΑΚΡΙΒΩΣ 12 χαρακτήρες.\n"
                                    "Επιτρέπονται: ελληνικά/λατινικά γράμματα και ψηφία.\n"
                                    "Ή αφήστε το πεδίο κενό αν δεν υπάρχει Προσωπικός Αριθμός.")
                self.personal_input.setFocus()
                return

        conn = Database.get_connection()
        if not conn:
            return

        try:
            cur = conn.cursor()

            # Parse πρωτόκολλο
            proto_match = re.match(r"(\d+)/(\d+)-(\d+)-(\d+)",
                                   self.protocol_input.text().strip())
            if not proto_match:
                QMessageBox.warning(self, "Σφάλμα",
                                    "Λάθος μορφή πρωτοκόλλου. Χρήση: ΝΝΝΝ/DD-MM-YYYY")
                return

            seq = int(proto_match.group(1))
            year = int(proto_match.group(4))

            # Ενημέρωση στη βάση
            cur.execute("""
                UPDATE health_certificates SET
                    protocol_number = ?,
                    proto_year = ?,
                    proto_seq = ?,
                    hc_first_name = ?,
                    hc_last_name = ?,
                    hc_father_name = ?,
                    hc_amka = ?,
                    hc_personal_number = ?,
                    id_number = ?,
                    passport_number = ?,
                    certificate_date = ?,
                    residence = ?,
                    work_type = ?,
                    comments = ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ? AND doctor_id = ?
            """, (
                self.protocol_input.text().strip(),
                year,
                seq,
                self.firstname_input.text().strip(),
                self.lastname_input.text().strip(),
                self.fathername_input.text().strip() or None,
                amka or None,
                personal_number or None,
                self.id_input.text().strip() or None,
                self.passport_input.text().strip() or None,
                self.date_input.date().toPyDate().isoformat(),
                self.residence_input.text().strip(),
                self.work_input.text().strip(),
                self.comments_input.toPlainText().strip() or None,
                cert_id,
                self.current_user_id
            ))

            if cur.rowcount == 0:
                QMessageBox.warning(self, "Προσοχή",
                                    "Το πιστοποιητικό δεν βρέθηκε ή δεν έχετε δικαίωμα ενημέρωσης.")
                return

            conn.commit()
            self.show_update_summary(cert_id)
            self.load_certificates()
            self.select_certificate_by_id(cert_id)

        except sqlite3.IntegrityError as e:
            conn.rollback()
            error_msg = str(e)

            if "UNIQUE constraint failed: health_certificates.hc_amka" in error_msg:
                QMessageBox.critical(self, "Σφάλμα",
                                     f"Ο ΑΜΚΑ '{amka}' υπάρχει ήδη σε άλλο πιστοποιητικό.\n"
                                     f"Παρακαλώ χρησιμοποιήστε διαφορετικό ΑΜΚΑ.")
            elif "UNIQUE constraint failed: health_certificates.hc_personal_number" in error_msg:
                QMessageBox.critical(self, "Σφάλμα",
                                     f"Ο Προσωπικός Αριθμός '{personal_number}' υπάρχει ήδη σε άλλο πιστοποιητικό.\n"
                                     f"Παρακαλώ χρησιμοποιήστε διαφορετικό Προσωπικό Αριθμό.")
            else:
                QMessageBox.critical(self, "Σφάλμα", f"Σφάλμα βάσης δεδομένων: {e}")
        except Exception as e:
            conn.rollback()
            QMessageBox.critical(self, "Σφάλμα", f"Σφάλμα ενημέρωσης: {e}")
            import traceback
            traceback.print_exc()

    def show_update_summary(self, cert_id):
        """Εμφανίζει σύνοψη των αλλαγών μετά την ενημέρωση."""
        # Βρες το πρωτόκολλο για το μήνυμα
        conn = Database.get_connection()
        if not conn:
            return

        try:
            cur = conn.cursor()
            cur.execute("""
                SELECT protocol_number, hc_last_name, hc_first_name
                FROM health_certificates
                WHERE id = ? AND doctor_id = ?
            """, (cert_id, self.current_user_id))

            result = cur.fetchone()
            if result:
                QMessageBox.information(self, "Επιτυχία Ενημέρωσης",
                                        f"Το πιστοποιητικό ενημερώθηκε επιτυχώς!\n\n"
                                        f"🔹 Πρωτόκολλο: {result[0]}\n"
                                        f"🔹 Ασθενής: {result[1]} {result[2]}\n"
                                        f"🔹 Ημερομηνία: {datetime.now().strftime('%d/%m/%Y %H:%M')}")

        except Exception as e:
            print(f"Σφάλμα σύνοψης ενημέρωσης: {e}")

    def select_certificate_by_id(self, cert_id):
        """Επιλέγει ένα πιστοποιητικό στον πίνακα με βάση το ID."""
        for i in range(self.certificates_table.topLevelItemCount()):
            item = self.certificates_table.topLevelItem(i)
            if item.text(0) == cert_id:
                self.certificates_table.setCurrentItem(item)
                item.setSelected(True)
                self.certificates_table.scrollToItem(item)
                break

    def delete_certificate(self):
        """Διαγράφει επιλεγμένο πιστοποιητικό."""
        selected = self.certificates_table.selectedItems()
        if not selected:
            QMessageBox.warning(self, "Προσοχή",
                                "Επιλέξτε ένα πιστοποιητικό για διαγραφή.")
            return

        cert_id = selected[0].text(0)

        reply = QMessageBox.question(
            self, "Επιβεβαίωση",
            "Είστε σίγουρος ότι θέλετε να διαγράψετε αυτό το πιστοποιητικό;\n"
            "Θα διαγραφούν και όλα τα συνημμένα αρχεία.",
            QMessageBox.Yes | QMessageBox.No
        )

        if reply != QMessageBox.Yes:
            return

        conn = Database.get_connection()
        if not conn:
            return

        try:
            cur = conn.cursor()

            # Διαγραφή συνημμένων πρώτα (λόγω foreign key)
            cur.execute("""
                DELETE FROM health_certificate_documents
                WHERE certificate_id = ? AND doctor_id = ?
            """, (cert_id, self.current_user_id))

            # Διαγραφή πιστοποιητικού
            cur.execute("""
                DELETE FROM health_certificates
                WHERE id = ? AND doctor_id = ?
            """, (cert_id, self.current_user_id))

            conn.commit()
            QMessageBox.information(self, "Επιτυχία",
                                    "Το πιστοποιητικό διαγράφτηκε.")

            self.clear_form()
            self.load_certificates()

        except Exception as e:
            conn.rollback()
            QMessageBox.critical(self, "Σφάλμα", f"Σφάλμα διαγραφής: {e}")
            import traceback
            traceback.print_exc()

    def manage_attachments(self):
        """Διαχείριση συνημμένων."""
        selected = self.certificates_table.selectedItems()
        if not selected:
            QMessageBox.warning(self, "Προσοχή", 
                              "Επιλέξτε ένα πιστοποιητικό.")
            return
        
        cert_id = selected[0].text(0)
        
        # Άνοιγμα διαλόγου διαχείρισης συνημμένων
        dialog = AttachmentDialog(self, cert_id)
        dialog.exec_()
        
        # Ανανέωση μετρητή συνημμένων
        self.update_attachments_count(cert_id)

    def clear_form(self):
        """Καθαρισμός όλων των πεδίων."""
        print("DEBUG: Αρχίζει clear_form")

        # Λίστα με όλα τα πεδία
        field_names = [
            'lastname_input', 'firstname_input', 'fathername_input',
            'amka_input', 'personal_input', 'id_input',
            'passport_input', 'residence_input', 'work_input'
        ]

        for field_name in field_names:
            if hasattr(self, field_name):
                field = getattr(self, field_name)
                field.clear()
                print(f"DEBUG: Καθάρισε το {field_name}")

        if hasattr(self, 'comments_input'):
            self.comments_input.clear()
            print("DEBUG: Καθάρισε comments_input")

        if hasattr(self, 'date_input'):
            self.date_input.setDate(QDate.currentDate())
            print(f"DEBUG: Ορίστηκε ημερομηνία: {QDate.currentDate().toString('dd/MM/yyyy')}")

        # Αυτόματη δημιουργία νέου πρωτοκόλλου
        if self.current_user_id:
            try:
                self.refresh_protocol()
                print("DEBUG: Ανανεώθηκε πρωτόκολλο")
            except Exception as e:
                print(f"DEBUG: Σφάλμα ανανέωσης πρωτοκόλλου: {e}")
        else:
            if hasattr(self, 'protocol_input'):
                self.protocol_input.clear()

        print("DEBUG: Τέλος clear_form")

    def on_certificate_select(self):
        """Όταν επιλέγεται πιστοποιητικό από τον πίνακα."""
        selected = self.certificates_table.selectedItems()
        if not selected:
            return

        # Πάρε το πρώτο επιλεγμένο αντικείμενο
        first_selected = selected[0]

        # Βρες τη γραμμή του επιλεγμένου κελιού
        if first_selected.parent() is None:  # Top-level item
            item = first_selected
        else:
            item = first_selected.parent()

        cert_id = item.text(0)  # ID είναι πάντα στη στήλη 0

        conn = Database.get_connection()
        if not conn:
            return

        try:
            cur = conn.cursor()

            cur.execute("""
                SELECT protocol_number, hc_first_name, hc_last_name,
                       hc_father_name, hc_amka, hc_personal_number,
                       id_number, passport_number, certificate_date,
                       residence, work_type, comments  -- 🔽 ΠΡΟΣΘΗΚΗ comments
                FROM health_certificates
                WHERE id = ? AND doctor_id = ?
            """, (cert_id, self.current_user_id))

            cert = cur.fetchone()
            if cert:
                # Συμπλήρωση όλων των πεδίων
                self.protocol_input.setText(cert[0] if cert[0] else "")
                self.firstname_input.setText(cert[1] if cert[1] else "")
                self.lastname_input.setText(cert[2] if cert[2] else "")
                self.fathername_input.setText(cert[3] if cert[3] else "")
                self.amka_input.setText(cert[4] if cert[4] else "")
                self.personal_input.setText(cert[5] if cert[5] else "")
                self.id_input.setText(cert[6] if cert[6] else "")  # Ταυτότητα
                self.passport_input.setText(cert[7] if cert[7] else "")  # Διαβατήριο

                # Χειρισμός ημερομηνίας
                date_value = cert[8]
                if date_value and hasattr(self, 'date_input'):
                    try:
                        if isinstance(date_value, str):
                            from datetime import datetime
                            try:
                                date_obj = datetime.strptime(date_value, "%Y-%m-%d")
                            except:
                                date_obj = datetime.strptime(date_value, "%Y-%m-%d %H:%M:%S")
                            self.date_input.setDate(QDate(date_obj.year,
                                                          date_obj.month,
                                                          date_obj.day))
                        else:
                            self.date_input.setDate(QDate(date_value.year,
                                                          date_value.month,
                                                          date_value.day))
                    except Exception as e:
                        print(f"DEBUG: Σφάλμα ημερομηνίας: {e}")
                        self.date_input.setDate(QDate.currentDate())

                self.residence_input.setText(cert[9] if cert[9] else "")
                self.work_input.setText(cert[10] if cert[10] else "")
                # 🔽 ΠΡΟΣΘΗΚΗ: Φόρτωση σχολίων
                self.comments_input.setPlainText(cert[11] if cert[11] else "")

        except Exception as e:
            print(f"Σφάλμα φόρτωσης στοιχείων: {e}")
            import traceback
            traceback.print_exc()

    def on_search(self):
        """Αναζήτηση πιστοποιητικών."""
        search_term = self.search_input.text().strip().lower()

        if not search_term:
            # Εμφάνιση όλων
            for i in range(self.certificates_table.topLevelItemCount()):
                item = self.certificates_table.topLevelItem(i)
                item.setHidden(False)
            return

        # Φιλτράρισμα
        for i in range(self.certificates_table.topLevelItemCount()):
            item = self.certificates_table.topLevelItem(i)

            # 🔽 ΕΝΗΜΕΡΩΣΗ: Προσθήκη σχολίων στην αναζήτηση
            matches = any([
                search_term in item.text(1).lower(),  # Πρωτόκολλο
                search_term in item.text(2).lower(),  # Επώνυμο
                search_term in item.text(3).lower(),  # Όνομα
                search_term in item.text(4).lower(),  # Πατρώνυμο
                search_term in item.text(5).lower(),  # ΑΜΚΑ
                search_term in item.text(6).lower(),  # Προσωπικός Αριθμός
                search_term in item.text(7).lower(),  # Ταυτότητα
                search_term in item.text(8).lower(),  # Διαβατήριο
                search_term in item.text(9).lower(),  # Ημερομηνία
                search_term in item.text(10).lower(),  # Κατοικία
                search_term in item.text(11).lower(),  # Εργασία
                search_term in item.text(12).lower(),  # Σχόλια 🔽 ΝΕΑ
            ])

            item.setHidden(not matches)

    def on_sort_changed(self):
        """Αλλαγή ταξινόμησης - ανανεώνει τον πίνακα."""
        # Απλώς ανανεώνουμε τον πίνακα - η πραγματική ταξινόμηση γίνεται στο SQL
        self.load_certificates()
    def export_to_excel(self):
        """Εξαγωγή σε Excel."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
            import os
            from datetime import datetime

            if self.certificates_table.topLevelItemCount() == 0:
                QMessageBox.warning(self, "Προσοχή", "Δεν υπάρχουν πιστοποιητικά για εξαγωγή.")
                return

            # Διάλογος για επιλογή αρχείου
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            default_filename = f"πιστοποιητικά_{timestamp}.xlsx"

            file_path, _ = QFileDialog.getSaveFileName(
                self, "Αποθήκευση Excel",
                os.path.join(os.path.dirname(os.path.abspath(__file__)), default_filename),
                "Excel Files (*.xlsx)"
            )

            if not file_path:
                return

            # Δημιουργία Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Πιστοποιητικά"

            # Επικεφαλίδες (με σχόλια)
            headers = [
                "Πρωτόκολλο", "Επώνυμο", "Όνομα", "Πατρώνυμο",
                "ΑΜΚΑ", "Προσωπικός Αριθμός", "Ταυτότητα", "Διαβατήριο",
                "Ημερομηνία", "Κατοικία", "Είδος Εργασίας", "Σχόλια"  # 🔽 ΠΡΟΣΘΗΚΗ
            ]

            # Προσθήκη επικεφαλίδων με στυλ
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, size=12)
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Προσθήκη δεδομένων
            conn = Database.get_connection()
            if not conn:
                QMessageBox.warning(self, "Σφάλμα", "Δεν ήταν δυνατή η σύνδεση με τη βάση δεδομένων.")
                return

            try:
                cur = conn.cursor()
                row = 2

                # Πάρε ΟΛΑ τα πιστοποιητικά του χρήστη (με σχόλια)
                cur.execute("""
                    SELECT protocol_number, hc_last_name, hc_first_name,
                           hc_father_name, hc_amka, hc_personal_number,
                           id_number, passport_number, certificate_date,
                           residence, work_type, comments  -- 🔽 ΠΡΟΣΘΗΚΗ
                    FROM health_certificates
                    WHERE doctor_id = ?
                    ORDER BY certificate_date DESC
                """, (self.current_user_id,))

                certificates = cur.fetchall()

                for cert in certificates:
                    # Μορφοποίηση ημερομηνίας
                    date_str = ""
                    if cert[8]:  # certificate_date
                        try:
                            if isinstance(cert[8], str):
                                try:
                                    date_obj = datetime.strptime(cert[8].split()[0], '%Y-%m-%d')
                                except:
                                    date_obj = datetime.strptime(cert[8], '%Y-%m-%d %H:%M:%S')
                                date_str = date_obj.strftime("%d/%m/%Y")
                            else:
                                date_str = cert[8].strftime("%d/%m/%Y")
                        except Exception as e:
                            print(f"DEBUG: Σφάλμα μορφοποίησης ημερομηνίας: {e}")
                            date_str = str(cert[8])[:10]

                    data = [
                        cert[0] if cert[0] else "",
                        cert[1] if cert[1] else "",
                        cert[2] if cert[2] else "",
                        cert[3] if cert[3] else "",
                        cert[4] if cert[4] else "",
                        cert[5] if cert[5] else "",
                        cert[6] if cert[6] else "",
                        cert[7] if cert[7] else "",
                        date_str,  # Μορφοποιημένη ημερομηνία
                        cert[9] if cert[9] else "",
                        cert[10] if cert[10] else "",
                        cert[11] if cert[11] else ""  # 🔽 Σχόλια
                    ]

                    # Προσθήκη στη γραμμή
                    for col, value in enumerate(data, 1):
                        ws.cell(row=row, column=col, value=value)

                    row += 1

            except Exception as e:
                print(f"Σφάλμα ανάκτησης δεδομένων: {e}")
                import traceback
                traceback.print_exc()

            # Αυτόματο προσαρμογή πλάτους στηλών
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Αποθήκευση
            wb.save(file_path)

            # Ερώτηση για άνοιγμα του αρχείου
            reply = QMessageBox.question(
                self, "✅ Επιτυχία",
                f"Εξήχθησαν {row - 2} πιστοποιητικά.\n\n"
                f"📄 Αρχείο: {os.path.basename(file_path)}\n\n"
                f"Θέλετε να ανοίξετε το αρχείο;",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes
            )

            if reply == QMessageBox.Yes:
                import subprocess
                if os.name == 'nt':  # Windows
                    os.startfile(file_path)
                elif os.name == 'posix':  # Linux/Mac
                    subprocess.run(['xdg-open', file_path])

        except ImportError:
            QMessageBox.critical(self, "Σφάλμα",
                                 "Το openpyxl δεν είναι εγκατεστημένο.\n"
                                 "Εγκαταστήστε με: pip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "Σφάλμα", f"Σφάλμα εξαγωγής: {e}")
            import traceback
            traceback.print_exc()

    def export_selected_to_excel(self):
        """Εξαγωγή μόνο των επιλεγμένων πιστοποιητικών σε Excel."""
        try:
            from openpyxl import Workbook
            import os
            from datetime import datetime

            # 🔽 ΑΛΛΑΓΗ: Χρήση selectedItems() για πολλαπλές επιλογές
            selected_items = self.certificates_table.selectedItems()

            if not selected_items:
                QMessageBox.warning(self, "Προσοχή", "Επιλέξτε πιστοποιητικά για εξαγωγή.")
                return

            # 🔽 ΑΛΛΑΓΗ: Λήψη μοναδικών γραμμών (όχι μεμονωμένων κελιών)
            # Το selectedItems() επιστρέφει κάθε επιλεγμένο κελί, οπότε χρειαζόμαστε μοναδικές γραμμές
            selected_rows = set()
            for item in selected_items:
                # Βρίσκουμε τη γονική γραμμή (αν είναι κελί)
                if item.parent() is None:  # Top-level item
                    row = self.certificates_table.indexOfTopLevelItem(item)
                else:
                    row = self.certificates_table.indexOfTopLevelItem(item.parent())
                selected_rows.add(row)

            if not selected_rows:
                QMessageBox.warning(self, "Προσοχή", "Δεν βρέθηκαν έγκυρες γραμμές.")
                return

            # Διαλογή των επιλεγμένων γραμμών
            selected_rows = sorted(selected_rows)

            # 🔽 ΑΛΛΑΓΗ: Λήψη των IDs από τις επιλεγμένες γραμμές
            cert_ids = []
            for row in selected_rows:
                item = self.certificates_table.topLevelItem(row)
                if item is not None:
                    cert_id = item.text(0)  # ID είναι στη στήλη 0 (κρυφή)
                    if cert_id and cert_id not in cert_ids:
                        cert_ids.append(cert_id)

            print(f"DEBUG: Επιλέχθηκαν {len(cert_ids)} πιστοποιητικά")

            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            default_filename = f"επιλεγμένα_πιστοποιητικά_{timestamp}.xlsx"

            file_path, _ = QFileDialog.getSaveFileName(
                self, "Αποθήκευση Excel",
                os.path.join(os.path.dirname(os.path.abspath(__file__)), default_filename),
                "Excel Files (*.xlsx)"
            )

            if not file_path:
                return

            # Δημιουργία Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Επιλεγμένα Πιστοποιητικά"

            # Επικεφαλίδες
            headers = [
                "Πρωτόκολλο", "Επώνυμο", "Όνομα", "Πατρώνυμο",
                "ΑΜΚΑ", "Προσωπικός Αριθμός", "Ταυτότητα", "Διαβατήριο",
                "Ημερομηνία", "Κατοικία", "Είδος Εργασίας", "Σχόλια"
            ]

            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
                ws.column_dimensions[chr(64 + col)].width = 20

            conn = Database.get_connection()
            if not conn:
                QMessageBox.warning(self, "Σφάλμα", "Δεν ήταν δυνατή η σύνδεση με τη βάση δεδομένων.")
                return

            row = 2
            exported_count = 0

            try:
                cur = conn.cursor()

                for cert_id in cert_ids:
                    cur.execute("""
                        SELECT protocol_number, hc_last_name, hc_first_name,
                               hc_father_name, hc_amka, hc_personal_number,
                               id_number, passport_number, certificate_date,
                               residence, work_type, comments
                        FROM health_certificates
                        WHERE id = ? AND doctor_id = ?
                    """, (cert_id, self.current_user_id))

                    cert = cur.fetchone()
                    if cert:
                        # Μορφοποίηση ημερομηνίας
                        date_str = ""
                        if cert[8]:  # certificate_date
                            try:
                                if isinstance(cert[8], str):
                                    try:
                                        date_obj = datetime.strptime(cert[8].split()[0], '%Y-%m-%d')
                                    except:
                                        date_obj = datetime.strptime(cert[8], '%Y-%m-%d %H:%M:%S')
                                    date_str = date_obj.strftime("%d/%m/%Y")
                                else:
                                    date_str = cert[8].strftime("%d/%m/%Y")
                            except Exception as e:
                                print(f"DEBUG: Σφάλμα μορφοποίησης ημερομηνίας: {e}")
                                date_str = str(cert[8])[:10]

                        data = [
                            cert[0] if cert[0] else "",
                            cert[1] if cert[1] else "",
                            cert[2] if cert[2] else "",
                            cert[3] if cert[3] else "",
                            cert[4] if cert[4] else "",
                            cert[5] if cert[5] else "",
                            cert[6] if cert[6] else "",
                            cert[7] if cert[7] else "",
                            date_str,  # Μορφοποιημένη ημερομηνία
                            cert[9] if cert[9] else "",
                            cert[10] if cert[10] else "",
                            cert[11] if cert[11] else ""
                        ]

                        for col, value in enumerate(data, 1):
                            ws.cell(row=row, column=col, value=value)

                        row += 1
                        exported_count += 1

                print(f"DEBUG: Εξήχθησαν {exported_count} πιστοποιητικά")  # Debug

            except Exception as e:
                print(f"Σφάλμα: {e}")
                import traceback
                traceback.print_exc()

            if exported_count == 0:
                QMessageBox.warning(self, "Προσοχή", "Δεν βρέθηκαν δεδομένα για εξαγωγή.")
                return

            # Αυτόματο προσαρμογή πλάτους
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(file_path)

            reply = QMessageBox.question(
                self, "✅ Επιτυχία",
                f"Εξήχθησαν {exported_count} πιστοποιητικά.\n\n"
                f"📄 Αρχείο: {os.path.basename(file_path)}\n\n"
                f"Θέλετε να ανοίξετε το αρχείο;",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes
            )

            if reply == QMessageBox.Yes:
                import subprocess
                if os.name == 'nt':
                    os.startfile(file_path)
                elif os.name == 'posix':
                    subprocess.run(['xdg-open', file_path])

        except ImportError:
            QMessageBox.critical(self, "Σφάλμα",
                                 "Το openpyxl δεν είναι εγκατεστημένο.\n"
                                 "Εγκαταστήστε με: pip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "Σφάλμα", f"Σφάλμα εξαγωγής: {e}")
            import traceback
            traceback.print_exc()

    def backup_to_csv(self):
        """Δημιουργεί backup όλων των δεδομένων σε CSV αρχείο."""
        import os
        import csv
        from datetime import datetime

        if self.certificates_table.topLevelItemCount() == 0:
            QMessageBox.warning(self, "Προσοχή", "Δεν υπάρχουν δεδομένα για backup.")
            return

        # Δημιουργία backup φακέλου αν δεν υπάρχει
        backup_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'backups')
        os.makedirs(backup_dir, exist_ok=True)

        # Δημιουργία ονόματος αρχείου
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        username = self.current_username if self.current_username else 'unknown'
        username = username.replace(' ', '_')
        filename = f"backup_{username}_{timestamp}.csv"
        file_path = os.path.join(backup_dir, filename)

        conn = Database.get_connection()
        if not conn:
            return

        try:
            cur = conn.cursor()

            # Ανάκτηση όλων των δεδομένων του τρέχοντος χρήστη
            cur.execute("""
                SELECT 
                    protocol_number,
                    hc_last_name,
                    hc_first_name,
                    hc_father_name,
                    hc_amka,
                    hc_personal_number,
                    id_number,
                    passport_number,
                    certificate_date,
                    residence,
                    work_type,
                    comments
                FROM health_certificates
                WHERE doctor_id = ?
                ORDER BY certificate_date DESC
            """, (self.current_user_id,))

            certificates = cur.fetchall()

            # Γράψιμο στο CSV
            with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)

                # Επικεφαλίδες
                headers = [
                    'Πρωτόκολλο', 'Επώνυμο', 'Όνομα', 'Πατρώνυμο',
                    'ΑΜΚΑ', 'Προσωπικός Αριθμός', 'Ταυτότητα', 'Διαβατήριο',
                    'Ημερομηνία Πιστοποιητικού', 'Κατοικία', 'Είδος Εργασίας', 'Σχόλια'
                ]
                writer.writerow(headers)

                # Δεδομένα
                count = 0
                for cert in certificates:
                    row = []
                    for i, value in enumerate(cert):
                        if value is None:
                            row.append('')
                        elif i == 8 and value:  # Ημερομηνία
                            try:
                                # Μετατροπή ημερομηνίας σε ελληνική μορφή
                                if isinstance(value, str):
                                    # Δοκιμάζουμε διαφορετικές μορφές
                                    try:
                                        date_obj = datetime.strptime(value.split()[0], '%Y-%m-%d')
                                    except:
                                        date_obj = datetime.strptime(value, '%Y-%m-%d %H:%M:%S')
                                    row.append(date_obj.strftime('%d/%m/%Y'))
                                else:
                                    # Αν είναι datetime object
                                    row.append(value.strftime('%d/%m/%Y'))
                            except:
                                row.append(str(value))
                        else:
                            row.append(str(value))
                    writer.writerow(row)
                    count += 1

            # Μήνυμα επιτυχίας
            QMessageBox.information(
                self,
                "✅ Επιτυχία Backup",
                f"Το backup δημιουργήθηκε επιτυχώς!\n\n"
                f"📁 Φάκελος: {backup_dir}\n"
                f"📄 Αρχείο: {filename}\n"
                f"📊 Εγγραφές: {count}\n\n"
                "Το CSV αρχείο μπορεί να ανοίξει με:\n"
                "• Excel\n• LibreOffice Calc\n• Google Sheets"
            )

            # Ερώτηση για άνοιγμα του φακέλου
            reply = QMessageBox.question(
                self, "Άνοιγμα Φακέλου",
                "Θέλετε να ανοίξετε τον φάκελο με τα backup;",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )

            if reply == QMessageBox.Yes:
                import subprocess
                if os.name == 'nt':  # Windows
                    os.startfile(backup_dir)
                elif os.name == 'posix':  # Linux/Mac
                    subprocess.run(['xdg-open', backup_dir])

        except Exception as e:
            QMessageBox.critical(self, "Σφάλμα", f"Σφάλμα κατά τη δημιουργία backup: {str(e)[:200]}")
            import traceback
            traceback.print_exc()

    def export_to_pdf(self):
        """Εξαγωγή σε PDF με την προεπιλεγμένη φόρμα."""
        self.export_selected_health_certificates_to_pdf_simple_form()

    def export_selected_health_certificates_to_pdf_simple_form(self):
        """
        Εκτυπώνει τις επιλεγμένες εγγραφές της καρτέλας 'Πιστοποιητικά Υγείας'
        σε A4 PDF με την ακριβή μορφοποίηση όπως στο Word έγγραφο.
        """
        from PyQt5.QtWidgets import QFileDialog, QMessageBox
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        import os
        from datetime import datetime
        from reportlab.lib.colors import gray, black, Color
        from reportlab.lib.utils import simpleSplit

        items = self.certificates_table.selectedItems()
        if not items:
            QMessageBox.information(self, "Εκτύπωση", "Επιλέξτε πιστοποιητικά για εκτύπωση.")
            return

        # Επιλογή αρχείου εξόδου
        out_path, _ = QFileDialog.getSaveFileName(
            self, "Αποθήκευση PDF",
            os.path.join(Config.EXPORTS_DIR, f"πιστοποιητικά_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"),
            "PDF (*.pdf)"
        )
        if not out_path:
            return

        # Προσπάθεια για ελληνική γραμματοσειρά
        font_name = "Helvetica"
        try:
            if os.path.exists("DejaVuSans.ttf"):
                pdfmetrics.registerFont(TTFont("DejaVuSans", "DejaVuSans.ttf"))
                font_name = "DejaVuSans"
                # Καταχωρούμε και Bold εκδοχή αν υπάρχει
                if os.path.exists("DejaVuSans-Bold.ttf"):
                    pdfmetrics.registerFont(TTFont("DejaVuSans-Bold", "DejaVuSans-Bold.ttf"))
        except Exception as e:
            print(f"Σφάλμα φόρτωσης γραμματοσειράς: {e}")
            pass

        # Ρυθμίσεις σελίδας
        width, height = A4
        left = 50  # αριστερό περιθώριο (pt)
        top = height - 50  # αρχικό y (pt)
        line = 20  # μειωμένη απόσταση γραμμών για καλύτερη χωρητικότητα

        c = canvas.Canvas(out_path, pagesize=A4)

        for it in items:
            # Πάρε το ID από το tree
            cert_id = it.text(0)

            # Φέρε τα πλήρη στοιχεία από τη βάση
            conn = Database.get_connection()
            if not conn:
                continue

            try:
                cur = conn.cursor()
                cur.execute("""
                    SELECT protocol_number, hc_first_name, hc_last_name,
                           hc_father_name, hc_amka, hc_personal_number,
                           id_number, passport_number, certificate_date,
                           residence, work_type
                    FROM health_certificates
                    WHERE id = ? AND doctor_id = ?
                """, (cert_id, self.current_user_id))

                cert = cur.fetchone()
                if not cert:
                    continue

                # Εξαγωγή δεδομένων
                protocol = cert[0] if cert[0] else ""
                first = cert[1] if cert[1] else ""
                last = cert[2] if cert[2] else ""
                father = cert[3] if cert[3] else ""
                amka = cert[4] if cert[4] else ""
                personal = cert[5] if cert[5] else ""
                id_no = cert[6] if cert[6] else ""
                passport = cert[7] if cert[7] else ""
                address = cert[9] if cert[9] else ""

                # Ημερομηνία
                cdate = ""
                date_value = cert[8]
                if date_value:
                    try:
                        if isinstance(date_value, str):
                            try:
                                date_obj = datetime.strptime(date_value, "%Y-%m-%d")
                            except:
                                date_obj = datetime.strptime(date_value, "%Y-%m-%d %H:%M:%S")
                            cdate = date_obj.strftime("%d/%m/%Y")
                        else:
                            cdate = date_value.strftime("%d/%m/%Y")
                    except Exception:
                        cdate = str(date_value)[:10]

                # Λήψη της σημερινής ημερομηνίας για το πρωτόκολλο
                current_date = datetime.now()

                # Δημιουργία αριθμού πρωτοκόλλου
                if not protocol:
                    # Μορφή: YYYY/DD-MM-YYYY όπως ζήτησες
                    protocol = f"{current_date.year}/{current_date.day:02d}-{current_date.month:02d}-{current_date.year}"

            except Exception as e:
                print(f"Σφάλμα φόρτωσης πιστοποιητικού: {e}")
                continue

            y = top

            # Αριθμός Πρωτοκόλλου στην επάνω αριστερή γωνία
            c.setFont(font_name, 10)
            c.drawString(left, y, f"Αρ.Πρωτ: {protocol}")
            y -= line * 6

            # Τίτλος - ΠΙΣΤΟΠΟΙΗΤΙΚΟ ΥΓΕΙΑΣ (BOLD)
            if font_name == "DejaVuSans" and hasattr(pdfmetrics, 'getFont'):
                try:
                    c.setFont("DejaVuSans-Bold", 14)
                except:
                    c.setFont(font_name, 14)
            else:
                c.setFont("Helvetica-Bold", 14)

            title = "ΠΙΣΤΟΠΟΙΗΤΙΚΟ ΥΓΕΙΑΣ"
            title_width = c.stringWidth(title, c._fontname, 14)
            c.drawString((width - title_width) / 2, y, title)
            y -= line

            # Υπότιτλος
            c.setFont(font_name, 9)
            subtitle = "(εργαζομένων σε καταστήματα, εργαστήρια ή επιχειρήσεις υγειονομικού ενδιαφέροντος)"
            subtitle_width = c.stringWidth(subtitle, font_name, 9)
            c.drawString((width - subtitle_width) / 2, y, subtitle)
            y -= line * 2

            # --- ΠΕΔΙΑ ΜΕ ΠΟΛΥ ΑΠΑΛΕΣ ΚΑΙ ΛΕΠΤΕΣ ΥΠΟΓΡΑΜΜΙΣΕΙΣ ---
            c.setFont(font_name, 10)
            name_x = left + 110

            # Δημιουργία πολύ απαλού γκρι χρώματος (σχεδόν διαφανές)
            very_light_gray = Color(0.9, 0.9, 0.9)  # Πολύ απαλό γκρι

            # Ο/Η Επώνυμο
            c.drawString(left, y, "Ο/Η Επώνυμο :")
            c.setFont(font_name, 10)
            c.drawString(name_x, y, f"{last}")

            # Πολύ απαλή και λεπτή υπογράμμιση
            underline_length = 200
            c.setStrokeColor(very_light_gray)  # Πολύ απαλό γκρι
            c.setLineWidth(0.3)  # Πολύ λεπτή γραμμή
            c.line(name_x, y - 1, name_x + underline_length, y - 1)  # Κοντύτερα στο κείμενο
            y -= line

            # Όνομα
            c.setFont(font_name, 10)
            c.drawString(left, y, "Όνομα :")
            c.setFont(font_name, 10)
            c.drawString(name_x, y, f"{first}")
            c.setStrokeColor(very_light_gray)
            c.setLineWidth(0.3)
            c.line(name_x, y - 1, name_x + underline_length, y - 1)
            y -= line

            # Όνομα Πατρός
            c.setFont(font_name, 10)
            c.drawString(left, y, "Όνομα Πατρός :")
            c.setFont(font_name, 10)
            c.drawString(name_x, y, f"{father}")
            c.setStrokeColor(very_light_gray)
            c.setLineWidth(0.3)
            c.line(name_x, y - 1, name_x + underline_length, y - 1)
            y -= line

            # Α. Τ.
            c.setFont(font_name, 10)
            c.drawString(left, y, "Α. Τ. :")
            c.setFont(font_name, 10)
            c.drawString(name_x, y, f"{id_no}")
            c.setStrokeColor(very_light_gray)
            c.setLineWidth(0.3)
            c.line(name_x, y - 1, name_x + underline_length, y - 1)
            y -= line

            # ή Αριθμός διαβατηρίου
            c.setFont(font_name, 10)
            c.drawString(left, y, "ή Αριθμός διαβατηρίου :")
            c.setFont(font_name, 10)
            c.drawString(name_x + 40, y, f"{passport}")
            c.setStrokeColor(very_light_gray)
            c.setLineWidth(0.3)
            c.line(name_x + 40, y - 1, name_x + underline_length, y - 1)
            y -= line

            # ή Α.Μ.Κ.Α.
            c.setFont(font_name, 10)
            c.drawString(left, y, "ή Α.Μ.Κ.Α. :")
            c.setFont(font_name, 10)
            c.drawString(name_x, y, f"{amka}")
            c.setStrokeColor(very_light_gray)
            c.setLineWidth(0.3)
            c.line(name_x, y - 1, name_x + underline_length, y - 1)
            y -= line
            
            # Δ/νση.Κατοικίας
            c.setFont(font_name, 10)
            c.drawString(left, y, "Δ/νση.Κατοικίας :")
            c.setFont(font_name, 10)
            c.drawString(name_x + 30, y, f"{address}")
            c.setStrokeColor(very_light_gray)
            c.setLineWidth(0.3)
            c.line(name_x + 30, y - 1, name_x + underline_length, y - 1)
            y -= line * 2.5  # Λίγο λιγότερο χώρο

            # Επαναφορά χρώματος για υπόλοιπο κείμενο
            c.setStrokeColor(black)
            c.setLineWidth(1)

            # Υποβλήθηκε σε Ιατρική κλινική εξέταση...
            c.setFont(font_name, 9)
            text = "Υποβλήθηκε σε Ιατρική κλινική εξέταση και στις παρακλινικές εξετάσεις"
            text_width = c.stringWidth(text, font_name, 9)
            c.drawString((width - text_width) / 5.5, y, text)
            y -= line

            # Ακτινογραφία Θώρακος (ΔΙΟΡΘΩΜΕΝΟ από "υπογραφή Θώρακος")
            exam_text = "Ακτινογραφία Θώρακος: □ Ναι Καλλιέργεια *, Παρασιτολογική κοπράνων (για χειριστές τροφίμων **) □ Ναι □ Όχι"
            text_width = c.stringWidth(exam_text, font_name, 9)
            c.drawString((width - text_width) / 1.5, y, exam_text)
            y -= line * 1

            # Κύριο κείμενο
            c.setFont(font_name, 9)
            main_text = "Βρέθηκε να μην πάσχει από κάποιο λοιμώδες ή μεταδοτικό νόσημα για τη Δημόσια Υγεία και μπορεί να εργαστεί σε κατάστημα, εργαστήριο ή επιχείρηση υγειονομικού ενδιαφέροντος."

            # Βελτιωμένη συνάρτηση για wrap με καλύτερο έλεγχο
            def draw_wrapped_text(text, x, y, max_width, font_name, font_size, line_spacing=None):
                if line_spacing is None:
                    line_spacing = line * 0.9  # Συντομότερο line spacing για υποσημειώσεις

                # Χρήση της ενσωματωμένης συνάρτησης simpleSplit για καλύτερο wrap
                c.setFont(font_name, font_size)
                lines = simpleSplit(text, font_name, font_size, max_width)

                # Σχεδίαση όλων των γραμμών
                for line_text in lines:
                    c.drawString(x, y, line_text)
                    y -= line_spacing

                return y, len(lines)

            # Υπολογισμός μέγιστου πλάτους για wrap
            max_text_width = width - 2 * left

            # Σχεδίαση του κύριου κειμένου
            new_y, lines_used = draw_wrapped_text(main_text, left, y, max_text_width, font_name, 10)
            y = new_y - line * 0.5  # Λίγο λιγότερο χώρο

            # Διάρκεια ισχύος
            validity_text = "Το παρόν ισχύει για πέντε (5) χρόνια, πλην των εργαζομένων σε παιδικούς, βρεφονηπιακούς σταθμούς κ.λ.π. που ισχύει για 2 χρόνια ***"

            new_y, lines_used2 = draw_wrapped_text(validity_text, left, y, max_text_width, font_name, 10)
            y = new_y - line * 2.1

            # ΗΜΕΡΟΜΗΝΙΑ & ΙΑΤΡΟΣ (δεξιά στοιχισμένα)
            c.setFont(font_name, 10)
            date_text = "ΗΜΕΡΟΜΗΝΙΑ :      /      / 20"
            date_width = c.stringWidth(date_text, font_name, 10)
            c.drawString(width - left - date_width, y, date_text)
            y -= line * 2.8

            doctor_text = "Ο ΙΑΤΡΟΣ"
            doctor_width = c.stringWidth(doctor_text, font_name, 10)
            c.drawString(width - left - doctor_width, y, doctor_text)
            y -= line * 0.7

            signature_text = "(υπογραφή, σφραγίδα)"
            signature_width = c.stringWidth(signature_text, font_name, 10)
            c.drawString(width - left - signature_width, y, signature_text)
            y -= line * 2.5

            # ΥΠΟΣΗΜΕΙΩΣΕΙΣ - ΠΟΙΟ ΣΥΜΑΖΕΜΕΝΕΣ ΚΑΙ ΜΕ ΜΙΚΡΟΤΕΡΑ ΓΡΑΜΜΑΤΑ
            c.setFont(font_name, 7)  # Μικρότερα γράμματα για υποσημειώσεις

            # Υποσημείωση 1
            footnote1 = "* Για salmonella ."
            c.drawString(left, y, footnote1)
            y -= line * 0.6  # Πολύ μικρή απόσταση

            # Υποσημείωση 2
            footnote2 = "** (Χειριστής τροφίμων είναι το άτομο που απευθείας χειρίζεται συσκευασμένα ή μη συσκευασμένα τρόφιμα, εξοπλισμό και εργαλεία για τρόφιμα ή επιφάνειες που έρχονται σε επαφή με τα τρόφιμα)."

            # Πιο συμπαγές wrap για υποσημείωση 2
            footnote2_width = width - 2 * left - 20  # Λίγο μικρότερο πλάτος για indent
            new_y, _ = draw_wrapped_text(footnote2, left + 5, y, footnote2_width, font_name, 7, line_spacing=line * 0.6)
            y = new_y - line * 0.3

            # Υποσημείωση 3
            footnote3 = "*** Τα Πιστοποιητικά υγείας του προσωπικού θα φυλάσσονται στις Επιχειρήσεις υγειονομικού ενδιαφέροντος και θα επιδεικνύονται όταν ζητούνται, στα αρμόδια όργανα υγειονομικού ελέγχου. Θα ανανεώνονται μετά τη λήξη τους επαναλαμβάνοντας όλες τις εξετάσεις (Ιατρική κλινική εξέταση)"

            # Πιο συμπαγές wrap για υποσημείωση 3
            new_y, _ = draw_wrapped_text(footnote3, left + 5, y, footnote2_width, font_name, 7, line_spacing=line * 0.6)

            # Έλεγχος αν χωράνε όλες οι υποσημειώσεις στην ίδια σελίδα
            if new_y < 50:  # Αν φτάνουμε πολύ κοντά στο τέλος της σελίδας
                # Αφήνουμε λίγο περιθώριο και συνεχίζουμε στην επόμενη σελίδα
                c.showPage()
                y = top - line * 5  # Ξεκινάμε από πιο ψηλά στην επόμενη σελίδα

            # Επόμενη σελίδα για επόμενο πιστοποιητικό
            c.showPage()

        c.save()
        QMessageBox.information(self, "PDF", f"Η εκτύπωση ολοκληρώθηκε: {out_path}")
            
    def print_certificate_form(self):
        """Εκτύπωση πιστοποιητικού σε προεπιλεγμένη φόρμα (PDF)."""
        self.export_selected_health_certificates_to_pdf_simple_form()

    def export_selected_health_certificates_to_pdf_on_form(self):
        """Εκτύπωση σε φόρμα με background εικόνα."""
        from PyQt5.QtWidgets import QFileDialog, QMessageBox
        from PyQt5.QtPrintSupport import QPrinter
        from PyQt5.QtGui import QPainter, QFont, QImage, QPen, QColor
        from PyQt5.QtCore import Qt, QRect
        import os
        from datetime import datetime  # 🔽 ΠΡΟΣΘΗΚΗ: Εισαγωγή datetime

        items = self.certificates_table.selectedItems()
        if not items:
            QMessageBox.information(self, "Εκτύπωση", "Δεν έχετε επιλέξει γραμμές.")
            return

        # Φόρμα (εικόνα Α4)
        template_path, _ = QFileDialog.getOpenFileName(self, "Επιλέξτε εικόνα φόρμας (A4)", "",
                                                       "Images (*.png *.jpg *.jpeg)")
        if not template_path:
            return
        bg = QImage(template_path)
        if bg.isNull():
            QMessageBox.critical(self, "Σφάλμα", "Αποτυχία φόρτωσης εικόνας φόρμας.")
            return

        out_path, _ = QFileDialog.getSaveFileName(
            self, "Αποθήκευση PDF",
            os.path.join(Config.EXPORTS_DIR, f"πιστοποιητικά_σε_φόρμα_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"),
            "PDF (*.pdf)"
        )
        if not out_path:
            return

        printer = QPrinter(QPrinter.HighResolution)
        printer.setOutputFormat(QPrinter.PdfFormat)
        printer.setOutputFileName(out_path)
        printer.setPaperSize(QPrinter.A4)
        printer.setPageMargins(10, 10, 10, 10, QPrinter.Millimeter)

        def mm(v: float) -> int:
            # mm -> pixels (int)
            return int(round(v * printer.resolution() / 25.4))

        # Μετατόπιση προς τα αριστερά
        SHIFT_X = -13.0
        SHIFT_Y = -10.0

        def mmo(x_mm: float, y_mm: float):
            # apply global shifts (σε mm) και γύρνα σε int px
            return mm(x_mm + SHIFT_X), mm(y_mm + SHIFT_Y)

        # Θέσεις (mm)
        X_LABEL, X_VALUE = 25.0, 90.0
        Y = {
            "proto": 25.0,
            "title": 55.0,
            "subtitle": 63.0,
            "lname": 80.0,
            "fname": 90.0,
            "father": 100.0,
            "amka": 110.0,
            "pers": 120.0,
            "id": 130.0,
            "pass": 140.0,
            "addr": 150.0,
        }

        p = QPainter(printer)
        try:
            for i, it in enumerate(items):
                if i > 0:
                    printer.newPage()

                page_rect = printer.pageRect()  # QRect
                p.drawImage(page_rect, bg)

                # Πάρε τα δεδομένα από τη βάση
                cert_id = it.text(0)
                conn = Database.get_connection()
                if conn:
                    try:
                        cur = conn.cursor()
                        cur.execute("""
                            SELECT protocol_number, hc_first_name, hc_last_name,
                                   hc_father_name, hc_amka, hc_personal_number,
                                   id_number, passport_number, residence
                            FROM health_certificates
                            WHERE id = ? AND doctor_id = ?
                        """, (cert_id, self.current_user_id))

                        cert = cur.fetchone()
                        if cert:
                            proto = cert[0] if cert[0] else ""
                            first = cert[1] if cert[1] else ""
                            last = cert[2] if cert[2] else ""
                            father = cert[3] if cert[3] else ""
                            amka = cert[4] if cert[4] else ""
                            personal = cert[5] if cert[5] else ""
                            id_no = cert[6] if cert[6] else ""
                            passport = cert[7] if cert[7] else ""
                            address = cert[8] if cert[8] else ""

                            p.setRenderHint(QPainter.Antialiasing, True)
                            p.setPen(QPen(QColor(0, 0, 0)))
                            p.setFont(QFont("DejaVu Sans", 12))

                            # Αρ.Πρωτ: κολλητά
                            x, y = mmo(X_LABEL, Y["proto"])
                            p.drawText(x, y, f"Αρ.Πρωτ: {proto}")

                            # Τίτλος στο κέντρο
                            title_y = mmo(0, Y["title"])[1]
                            title_rect = QRect(page_rect.left(), title_y - mm(5), page_rect.width(), mm(10))
                            p.setFont(QFont("DejaVu Sans", 15, QFont.Bold))
                            p.drawText(title_rect, Qt.AlignHCenter | Qt.AlignVCenter, "ΠΙΣΤΟΠΟΙΗΤΙΚΟ ΥΓΕΙΑΣ")

                            # Υπότιτλος στο κέντρο
                            sub_y = mmo(0, Y["subtitle"])[1]
                            sub_rect = QRect(page_rect.left(), sub_y - mm(4), page_rect.width(), mm(8))
                            p.setFont(QFont("DejaVu Sans", 10))
                            p.drawText(sub_rect, Qt.AlignHCenter | Qt.AlignVCenter,
                                       "(εργαζομένων σε καταστήματα, εργαστήρια ή επιχειρήσεις υγειονομικού ενδιαφέροντος)")

                            # Επιστροφή σε κανονική γραμματοσειρά για τα πεδία
                            p.setFont(QFont("DejaVu Sans", 12))

                            # Γραμμή-γραμμή
                            x, y = mmo(X_LABEL, Y["lname"]);
                            p.drawText(x, y, f"Ο/Η Επώνυμο : {last}")
                            x, y = mmo(X_LABEL, Y["fname"]);
                            p.drawText(x, y, f"Όνομα : {first}")
                            x, y = mmo(X_LABEL, Y["father"]);
                            p.drawText(x, y, f"Όνομα Πατρός : {father}")
                            x, y = mmo(X_LABEL, Y["amka"]);
                            p.drawText(x, y, f"Α.Μ.Κ.Α. : {amka}")
                            x, y = mmo(X_LABEL, Y["pers"]);
                            p.drawText(x, y, f"Προσωπικός αριθμός : {personal}")
                            x, y = mmo(X_LABEL, Y["id"]);
                            p.drawText(x, y, f"Α. Τ. : {id_no}")
                            x, y = mmo(X_LABEL, Y["pass"]);
                            p.drawText(x, y, f"ή Αριθμός διαβατηρίου : {passport}")
                            x, y = mmo(X_LABEL, Y["addr"]);
                            p.drawText(x, y, f"Δ/νση.Κατοικίας : {address}")

                    except Exception as e:
                        print(f"Σφάλμα: {e}")

        finally:
            p.end()

        QMessageBox.information(self, "PDF", f"Αποθηκεύτηκε: {out_path}")
                
class AttachmentDialog(QDialog):
    """Διάλογος διαχείρισης συνημμένων."""

    def __init__(self, parent, certificate_id):
        super().__init__(parent)
        self.parent = parent
        self.certificate_id = certificate_id

        self.setWindowTitle("Διαχείριση Συνημμένων")
        self.setMinimumSize(500, 400)

        # 🔽 ΠΡΟΣΘΗΚΗ: Δημιουργία φακέλου για συνημμένα
        self.create_attachments_directory()

        self.setup_ui()
        self.load_attachments()

    def create_attachments_directory(self):
        """Δημιουργεί τον φάκελο για τα συνημμένα."""
        import os
        if self.parent and self.parent.current_username:
            user_dir = os.path.join(Config.ATTACHMENTS_DIR,
                                    self.parent.current_username)
            cert_dir = os.path.join(user_dir, self.certificate_id)
            os.makedirs(cert_dir, exist_ok=True)

    def setup_ui(self):
        """Ρυθμίζει το UI του dialog."""
        layout = QVBoxLayout()
        self.setLayout(layout)
        
        # Λίστα αρχείων
        self.files_list = QListWidget()
        layout.addWidget(self.files_list)
        
        # Κουμπιά
        buttons_layout = QHBoxLayout()
        
        self.add_btn = QPushButton("➕ Προσθήκη")
        self.add_btn.clicked.connect(self.add_file)
        
        self.open_btn = QPushButton("👁️ Άνοιγμα")
        self.open_btn.clicked.connect(self.open_file)
        
        self.delete_btn = QPushButton("🗑️ Διαγραφή")
        self.delete_btn.clicked.connect(self.delete_file)
        
        self.close_btn = QPushButton("Κλείσιμο")
        self.close_btn.clicked.connect(self.accept)
        
        buttons_layout.addWidget(self.add_btn)
        buttons_layout.addWidget(self.open_btn)
        buttons_layout.addWidget(self.delete_btn)
        buttons_layout.addStretch()
        buttons_layout.addWidget(self.close_btn)
        
        layout.addLayout(buttons_layout)

    def load_attachments(self):
        """Φορτώνει τα συνημμένα από τη βάση."""
        self.files_list.clear()

        conn = Database.get_connection()
        if not conn:
            return

        try:
            cur = conn.cursor()

            cur.execute("""
                SELECT id, file_name, file_size, upload_date
                FROM health_certificate_documents
                WHERE certificate_id = ? AND doctor_id = ?
                ORDER BY upload_date DESC
            """, (self.certificate_id, self.parent.current_user_id))

            files = cur.fetchall()

            for file_id, file_name, file_size, upload_date in files:
                size_str = self.format_file_size(file_size) if file_size else "0 B"

                # 🔽 ΔΙΟΡΘΩΣΗ: Χειρισμός ημερομηνίας από string
                date_str = ""
                if upload_date:
                    try:
                        if isinstance(upload_date, str):
                            # Αν είναι string, κάνε parse
                            from datetime import datetime
                            date_obj = datetime.strptime(upload_date, "%Y-%m-%d %H:%M:%S")
                            date_str = date_obj.strftime("%d/%m/%Y %H:%M")
                        else:
                            # Αν είναι datetime object
                            date_str = upload_date.strftime("%d/%m/%Y %H:%M")
                    except Exception as e:
                        print(f"Σφάλμα μορφοποίησης ημερομηνίας: {e}")
                        date_str = str(upload_date)[:19]  # Πάρε τα πρώτα 19 χαρακτήρες

                item_text = f"{file_name} ({size_str})"
                if date_str:
                    item_text += f" - {date_str}"

                self.files_list.addItem(item_text)
                self.files_list.item(self.files_list.count() - 1).setData(Qt.UserRole, file_id)

        except Exception as e:
            QMessageBox.critical(self, "Σφάλμα", f"Σφάλμα φόρτωσης: {e}")
            import traceback
            traceback.print_exc()

    def format_file_size(self, size_bytes):
        """Μορφοποιεί το μέγεθος αρχείου."""
        if size_bytes is None:
            return "0 B"

        units = ['B', 'KB', 'MB', 'GB']
        size = float(size_bytes)
        unit_index = 0

        while size >= 1024 and unit_index < len(units) - 1:
            size /= 1024
            unit_index += 1

        return f"{size:.1f} {units[unit_index]}"

    def add_file(self):
        """Προσθήκη νέου αρχείου."""
        file_dialog = QFileDialog()
        file_dialog.setFileMode(QFileDialog.ExistingFiles)

        if file_dialog.exec_():
            files = file_dialog.selectedFiles()

            for file_path in files:
                self.upload_file(file_path)

    def upload_file(self, file_path):
        """Ανεβάζει αρχείο στον server."""
        import shutil
        import os

        # Δημιουργία φακέλου
        user_dir = os.path.join(Config.ATTACHMENTS_DIR,
                                self.parent.current_username)
        cert_dir = os.path.join(user_dir, self.certificate_id)
        os.makedirs(cert_dir, exist_ok=True)

        try:
            # Αντιγραφή αρχείου
            file_name = os.path.basename(file_path)
            dest_path = os.path.join(cert_dir, file_name)
            shutil.copy2(file_path, dest_path)

            # Πληροφορίες αρχείου
            file_size = os.path.getsize(file_path)

            # Αποθήκευση στη βάση
            conn = Database.get_connection()
            if conn:
                try:
                    cur = conn.cursor()  # 🔽 ΑΛΛΑΓΗ: Χωρίς with

                    cur.execute("""
                        INSERT INTO health_certificate_documents
                        (certificate_id, file_name, file_path, file_size, doctor_id)
                        VALUES (?, ?, ?, ?, ?)
                    """, (
                        self.certificate_id,
                        file_name,
                        dest_path,
                        file_size,
                        self.parent.current_user_id
                    ))
                    conn.commit()

                except Exception as e:
                    conn.rollback()
                    raise e

            # Ανανέωση λίστας
            self.load_attachments()

            QMessageBox.information(self, "Επιτυχία",
                                    f"Το αρχείο '{file_name}' προστέθηκε.")

        except Exception as e:
            QMessageBox.critical(self, "Σφάλμα", f"Σφάλμα ανεβάσματος: {e}")
            import traceback
            traceback.print_exc()

    def open_file(self):
        """Ανοίγει το επιλεγμένο αρχείο."""
        selected = self.files_list.selectedItems()
        if not selected:
            QMessageBox.warning(self, "Προσοχή", "Επιλέξτε ένα αρχείο.")
            return

        # Βρες τη διαδρομή του αρχείου από τη βάση
        file_id = selected[0].data(Qt.UserRole)

        conn = Database.get_connection()
        if not conn:
            return

        try:
            cur = conn.cursor()  # 🔽 ΑΛΛΑΓΗ: Χωρίς with

            cur.execute("""
                SELECT file_path FROM health_certificate_documents
                WHERE id = ?
            """, (file_id,))

            result = cur.fetchone()
            if result and os.path.exists(result[0]):
                import subprocess
                subprocess.run(['start', '', result[0]], shell=True)
            else:
                QMessageBox.warning(self, "Προσοχή",
                                    "Το αρχείο δεν βρέθηκε.")

        except Exception as e:
            QMessageBox.critical(self, "Σφάλμα", f"Σφάλμα ανοίγματος: {e}")

    def delete_file(self):
        """Διαγράφει το επιλεγμένο αρχείο."""
        selected = self.files_list.selectedItems()
        if not selected:
            return

        reply = QMessageBox.question(
            self, "Επιβεβαίωση",
            "Είστε σίγουρος ότι θέλετε να διαγράψετε αυτό το αρχείο;",
            QMessageBox.Yes | QMessageBox.No
        )

        if reply != QMessageBox.Yes:
            return

        file_id = selected[0].data(Qt.UserRole)

        conn = Database.get_connection()
        if not conn:
            return

        try:
            cur = conn.cursor()  # 🔽 ΑΛΛΑΓΗ: Χωρίς with

            # Βρες τη διαδρομή για διαγραφή από το filesystem
            cur.execute("""
                SELECT file_path FROM health_certificate_documents
                WHERE id = ?
            """, (file_id,))

            result = cur.fetchone()

            # Διαγραφή από τη βάση
            cur.execute("""
                DELETE FROM health_certificate_documents
                WHERE id = ?
            """, (file_id,))

            # Διαγραφή από το filesystem
            if result and os.path.exists(result[0]):
                os.remove(result[0])

            conn.commit()

            # Ανανέωση
            self.load_attachments()

            QMessageBox.information(self, "Επιτυχία",
                                    "Το αρχείο διαγράφτηκε.")

        except Exception as e:
            conn.rollback()
            QMessageBox.critical(self, "Σφάλμα", f"Σφάλμα διαγραφής: {e}")
            import traceback
            traceback.print_exc()

class UserManagementDialog(QDialog):
    """Διάλογος διαχείρισης χρηστών."""

    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.setWindowTitle("👤 Διαχείριση Χρηστών")
        self.setMinimumSize(600, 400)
        self.setup_ui()
        self.load_users()

    def setup_ui(self):
        """Ρυθμίζει το UI του dialog."""
        layout = QVBoxLayout()
        self.setLayout(layout)

        # Λίστα χρηστών
        self.users_list = QTreeWidget()
        self.users_list.setColumnCount(5)
        self.users_list.setHeaderLabels(["ID", "Όνομα Χρήστη", "Ρόλος", "Κατάσταση", "Ημερομηνία Δημιουργίας"])
        self.users_list.setColumnHidden(0, True)  # Κρύβουμε το ID
        self.users_list.setAlternatingRowColors(True)

        # Αυτόματο προσαρμογή πλάτους
        self.users_list.header().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.users_list.header().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.users_list.header().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.users_list.header().setSectionResizeMode(4, QHeaderView.Stretch)

        layout.addWidget(self.users_list)

        # Κουμπιά
        buttons_layout = QHBoxLayout()

        self.add_user_btn = QPushButton("➕ Νέος Χρήστης")
        self.add_user_btn.clicked.connect(self.add_user)

        self.edit_user_btn = QPushButton("✏️ Επεξεργασία")
        self.edit_user_btn.clicked.connect(self.edit_user)

        self.delete_user_btn = QPushButton("🗑️ Διαγραφή")
        self.delete_user_btn.clicked.connect(self.delete_user)

        self.change_password_btn = QPushButton("🔑 Αλλαγή Κωδικού")
        self.change_password_btn.clicked.connect(self.change_password)

        self.close_btn = QPushButton("Κλείσιμο")
        self.close_btn.clicked.connect(self.accept)

        buttons_layout.addWidget(self.add_user_btn)
        buttons_layout.addWidget(self.edit_user_btn)
        buttons_layout.addWidget(self.delete_user_btn)
        buttons_layout.addWidget(self.change_password_btn)
        buttons_layout.addStretch()
        buttons_layout.addWidget(self.close_btn)

        layout.addLayout(buttons_layout)

    def load_users(self):
        """Φορτώνει τους χρήστες από τη βάση."""
        self.users_list.clear()

        conn = Database.get_connection()
        if not conn:
            return

        try:
            cur = conn.cursor()
            cur.execute("""
                SELECT id, username, role, is_active, created_at
                FROM users
                ORDER BY username
            """)

            users = cur.fetchall()

            for user in users:
                item = QTreeWidgetItem(self.users_list)
                item.setText(0, str(user[0]))  # ID
                item.setText(1, user[1] if user[1] else "")  # Username
                item.setText(2, user[2] if user[2] else "")  # Role
                item.setText(3, "Ενεργός" if user[3] else "Ανενεργός")  # Status

                # Μορφοποίηση ημερομηνίας
                date_str = user[4] if user[4] else ""
                if date_str:
                    try:
                        if isinstance(date_str, str):
                            from datetime import datetime
                            date_obj = datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
                            date_str = date_obj.strftime("%d/%m/%Y %H:%M")
                        else:
                            date_str = date_str.strftime("%d/%m/%Y %H:%M")
                    except Exception:
                        date_str = str(date_str)[:19]
                item.setText(4, date_str)

        except Exception as e:
            QMessageBox.critical(self, "Σφάλμα", f"Σφάλμα φόρτωσης χρηστών: {e}")

    def add_user(self):
        """Προσθέτει νέο χρήστη."""
        dialog = UserEditDialog(self, None)
        if dialog.exec_() == QDialog.Accepted:
            self.load_users()

    def edit_user(self):
        """Επεξεργάζεται υπάρχοντα χρήστη."""
        selected = self.users_list.selectedItems()
        if not selected:
            QMessageBox.warning(self, "Προσοχή", "Επιλέξτε έναν χρήστη.")
            return

        user_id = selected[0].text(0)
        dialog = UserEditDialog(self, user_id)
        if dialog.exec_() == QDialog.Accepted:
            self.load_users()

    def delete_user(self):
        """Διαγράφει χρήστη."""
        selected = self.users_list.selectedItems()
        if not selected:
            QMessageBox.warning(self, "Προσοχή", "Επιλέξτε έναν χρήστη.")
            return

        user_id = selected[0].text(0)
        username = selected[0].text(1)

        # Έλεγχος αν είναι ο προεπιλεγμένος admin
        conn = Database.get_connection()
        if conn:
            try:
                cur = conn.cursor()
                cur.execute("SELECT is_default_admin FROM users WHERE id = ?", (user_id,))
                result = cur.fetchone()
                is_default_admin = result[0] if result else 0

                if is_default_admin:
                    reply = QMessageBox.warning(
                        self, "Προσοχή",
                        "Διαγράφετε τον προεπιλεγμένο διαχειριστή!\n\n"
                        "Ο χρήστης 'admin' θα δημιουργηθεί ξανά κατά την επόμενη εκκίνηση του προγράμματος.\n\n"
                        "Θέλετε να συνεχίσετε;",
                        QMessageBox.Yes | QMessageBox.No,
                        QMessageBox.No
                    )
                    if reply != QMessageBox.Yes:
                        return

            except Exception as e:
                print(f"Σφάλμα ελέγχου προεπιλεγμένου admin: {e}")

        # Δεν επιτρέπουμε διαγραφή του τρέχοντος χρήστη
        if user_id == self.parent.current_user_id:
            QMessageBox.warning(self, "Προσοχή",
                                "Δεν μπορείτε να διαγράψετε τον τρέχοντα χρήστη.")
            return

        reply = QMessageBox.question(
            self, "Επιβεβαίωση",
            f"Είστε σίγουρος ότι θέλετε να διαγράψετε τον χρήστη '{username}'?",
            QMessageBox.Yes | QMessageBox.No
        )

        if reply != QMessageBox.Yes:
            return

        conn = Database.get_connection()
        if not conn:
            return

        try:
            cur = conn.cursor()
            cur.execute("DELETE FROM users WHERE id = ?", (user_id,))
            conn.commit()

            QMessageBox.information(self, "Επιτυχία", "Ο χρήστης διαγράφτηκε.")
            self.load_users()

        except Exception as e:
            conn.rollback()
            QMessageBox.critical(self, "Σφάλμα", f"Σφάλμα διαγραφής: {e}")

    def change_password(self):
        """Αλλάζει κωδικό χρήστη."""
        selected = self.users_list.selectedItems()
        if not selected:
            QMessageBox.warning(self, "Προσοχή", "Επιλέξτε έναν χρήστη.")
            return

        user_id = selected[0].text(0)
        username = selected[0].text(1)

        # Διάλογος αλλαγής κωδικού
        new_password, ok = QInputDialog.getText(
            self, "Αλλαγή Κωδικού",
            f"Νέος κωδικός για τον χρήστη '{username}':",
            QLineEdit.Password
        )

        if not ok or not new_password:
            return

        # Επιβεβαίωση κωδικού
        confirm_password, ok = QInputDialog.getText(
            self, "Επιβεβαίωση Κωδικού",
            "Επιβεβαίωση νέου κωδικού:",
            QLineEdit.Password
        )

        if not ok:
            return

        if new_password != confirm_password:
            QMessageBox.warning(self, "Σφάλμα", "Οι κωδικοί δεν ταιριάζουν.")
            return

        conn = Database.get_connection()
        if not conn:
            return

        try:
            cur = conn.cursor()
            cur.execute("""
                UPDATE users 
                SET password_hash = ?
                WHERE id = ?
            """, (new_password, user_id))

            conn.commit()
            QMessageBox.information(self, "Επιτυχία", "Ο κωδικός άλλαξε επιτυχώς.")

        except Exception as e:
            conn.rollback()
            QMessageBox.critical(self, "Σφάλμα", f"Σφάλμα αλλαγής κωδικού: {e}")

class UserEditDialog(QDialog):
    """Διάλογος επεξεργασίας/προσθήκης χρήστη."""

    def __init__(self, parent, user_id):
        super().__init__(parent)
        self.parent = parent
        self.user_id = user_id
        self.is_edit = user_id is not None

        self.setWindowTitle("✏️ Επεξεργασία Χρήστη" if self.is_edit else "➕ Νέος Χρήστης")
        self.setMinimumSize(400, 300)
        self.setup_ui()

        if self.is_edit:
            self.load_user_data()

    def setup_ui(self):
        """Ρυθμίζει το UI του dialog."""
        layout = QVBoxLayout()
        self.setLayout(layout)

        form_layout = QVBoxLayout()

        # Όνομα Χρήστη
        username_label = QLabel("Όνομα Χρήστη:")
        self.username_input = QLineEdit()
        self.username_input.setPlaceholderText("Εισάγετε όνομα χρήστη")

        form_layout.addWidget(username_label)
        form_layout.addWidget(self.username_input)

        # Κωδικός (μόνο για νέους χρήστες)
        if not self.is_edit:
            password_label = QLabel("Κωδικός:")
            self.password_input = QLineEdit()
            self.password_input.setEchoMode(QLineEdit.Password)
            self.password_input.setPlaceholderText("Εισάγετε κωδικό")

            confirm_label = QLabel("Επιβεβαίωση Κωδικού:")
            self.confirm_input = QLineEdit()
            self.confirm_input.setEchoMode(QLineEdit.Password)
            self.confirm_input.setPlaceholderText("Επιβεβαιώστε τον κωδικό")

            form_layout.addWidget(password_label)
            form_layout.addWidget(self.password_input)
            form_layout.addWidget(confirm_label)
            form_layout.addWidget(self.confirm_input)

        # Ρόλος
        role_label = QLabel("Ρόλος:")
        self.role_combo = QComboBox()
        self.role_combo.addItems(["doctor", "admin"])

        form_layout.addWidget(role_label)
        form_layout.addWidget(self.role_combo)

        # Κατάσταση
        status_label = QLabel("Κατάσταση:")
        self.status_combo = QComboBox()
        self.status_combo.addItems(["Ενεργός", "Ανενεργός"])

        form_layout.addWidget(status_label)
        form_layout.addWidget(self.status_combo)

        layout.addLayout(form_layout)

        # Κουμπιά
        buttons_layout = QHBoxLayout()

        self.save_btn = QPushButton("💾 Αποθήκευση")
        self.save_btn.clicked.connect(self.save_user)

        self.cancel_btn = QPushButton("❌ Ακύρωση")
        self.cancel_btn.clicked.connect(self.reject)

        buttons_layout.addStretch()
        buttons_layout.addWidget(self.save_btn)
        buttons_layout.addWidget(self.cancel_btn)

        layout.addLayout(buttons_layout)

    def load_user_data(self):
        """Φορτώνει τα δεδομένα του χρήστη."""
        conn = Database.get_connection()
        if not conn:
            return

        try:
            cur = conn.cursor()
            cur.execute("""
                SELECT username, role, is_active
                FROM users
                WHERE id = ?
            """, (self.user_id,))

            user = cur.fetchone()
            if user:
                self.username_input.setText(user[0] if user[0] else "")
                self.role_combo.setCurrentText(user[1] if user[1] else "doctor")
                self.status_combo.setCurrentIndex(0 if user[2] else 1)

        except Exception as e:
            print(f"Σφάλμα φόρτωσης χρήστη: {e}")

    def save_user(self):
        """Αποθηκεύει τον χρήστη."""
        username = self.username_input.text().strip()
        role = self.role_combo.currentText()
        is_active = self.status_combo.currentText() == "Ενεργός"

        # Έλεγχος ονόματος χρήστη
        if not username:
            QMessageBox.warning(self, "Σφάλμα", "Το όνομα χρήστη είναι υποχρεωτικό.")
            self.username_input.setFocus()
            return

        # Για νέους χρήστες, έλεγχος κωδικού
        if not self.is_edit:
            password = self.password_input.text().strip()
            confirm = self.confirm_input.text().strip()

            if not password:
                QMessageBox.warning(self, "Σφάλμα", "Ο κωδικός είναι υποχρεωτικός.")
                self.password_input.setFocus()
                return

            if password != confirm:
                QMessageBox.warning(self, "Σφάλμα", "Οι κωδικοί δεν ταιριάζουν.")
                self.password_input.setFocus()
                return

        conn = Database.get_connection()
        if not conn:
            return

        try:
            cur = conn.cursor()

            if self.is_edit:
                # Ενημέρωση υπάρχοντος χρήστη
                cur.execute("""
                    UPDATE users 
                    SET username = ?, role = ?, is_active = ?
                    WHERE id = ?
                """, (username, role, is_active, self.user_id))
            else:
                # Προσθήκη νέου χρήστη
                import uuid
                cur.execute("""
                    INSERT INTO users (id, username, password_hash, role, is_active)
                    VALUES (?, ?, ?, ?, ?)
                """, (str(uuid.uuid4()), username, password, role, is_active))

            conn.commit()
            QMessageBox.information(self, "Επιτυχία",
                                    "Ο χρήστης αποθηκεύτηκε επιτυχώς.")
            self.accept()

        except Exception as e:
            conn.rollback()

            # Έλεγχος για διπλότυπο όνομα χρήστη
            if "UNIQUE constraint failed" in str(e) and "username" in str(e):
                QMessageBox.critical(self, "Σφάλμα",
                                     f"Το όνομα χρήστη '{username}' υπάρχει ήδη.")
                self.username_input.setFocus()
            else:
                QMessageBox.critical(self, "Σφάλμα", f"Σφάλμα αποθήκευσης: {e}")
        
def main():
    app = QApplication(sys.argv)

    # Ορισμός ελληνικών γραμματοσειρών
    font = app.font()
    font.setPointSize(10)
    app.setFont(font)

    window = HealthCertificateApp()
    window.show()

    sys.exit(app.exec_())


if __name__ == "__main__":
    main()